"""
Excel Compiler Application
Version: 3.0
Auteur: GOUNOU N'GOBI Chabi Zimé (Data Manager, Data Analyst)
Améliorations: Juin 2025

Application pour compiler plusieurs fichiers Excel en un seul fichier avec diverses options de formatage.
Nouvelles fonctionnalités:
- Prévisualisation des données
- Internationalisation complète
- Options de format de date
"""

import sys
import os

import pandas as pd
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QWidget, QTabWidget,
    QGroupBox, QLabel, QSpinBox, QCheckBox, QLineEdit, QPushButton, QListWidget,
    QProgressBar, QMessageBox, QFileDialog, QListWidgetItem, QStyle, QProgressDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QScrollArea, QDialog, QComboBox,
    QStyledItemDelegate, QStyleOptionButton, QGridLayout, QRadioButton,
    QButtonGroup, QSplitter, QToolBar, QMenu, QMenuBar
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer, QSize, QSettings, QTranslator, QLocale
from PyQt6.QtGui import QIcon, QFont, QPalette, QColor, QBrush, QKeySequence, QAction
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, numbers
from openpyxl.utils import get_column_letter
import openpyxl
import logging
import csv
import json
import unittest
from datetime import datetime
from typing import List, Tuple, Dict, Optional, Any, Union, Set
import traceback


def resource_path(relative_path):
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, relative_path)
        return os.path.join(os.path.abspath("."), relative_path)


# Constants for styles
COLORS = {
    "PRIMARY": "2e7d32",
    "PRIMARY_DARK": "2e7d32",
    "PRIMARY_LIGHT": "4caf50",  
    "ACCENT": "c8e6c9",
    "BACKGROUND": "2e7d32",
    "LIGHT_TEXT": "FFFFFF",
    "DARK_TEXT": "212121",
    "BORDER": "e0e0e0",
    "WARNING": "f44336",
    "SUCCESS": "4CAF50",
    "INFO": "2196F3"
}

# Pour openpyxl, ajoutez FF au début pour l'opacité
EXCEL_COLORS = {
    "PRIMARY": "FF2e7d32",
    "PRIMARY_DARK": "FF2e7d32",
    "PRIMARY_LIGHT": "FF4caf50",
    "ACCENT": "FFc8e6c9",
    "BACKGROUND":"FF2e7d32",
    "LIGHT_TEXT": "FFFFFFFF",
    "DARK_TEXT": "FF212121",
    "BORDER": "FFe0e0e0",
    "WARNING": "FFf44336",
    "SUCCESS": "FF4CAF50",
    "INFO": "FF2196F3"
}

FONT_SIZES = {
    "SMALL": 9,
    "NORMAL": 10,
    "LARGE": 12,
    "HEADER": 14
}

# Formats de date disponibles
DATE_FORMATS = {
    "STANDARD": {"format": "yyyy-MM-dd", "code": "yyyy-mm-dd", "excel_format": "yyyy-mm-dd"},
    "FRENCH": {"format": "dd/MM/yyyy", "code": "dd/mm/yyyy", "excel_format": "dd/mm/yyyy"},
    "US": {"format": "MM/dd/yyyy", "code": "mm/dd/yyyy", "excel_format": "mm/dd/yyyy"},
    "DATETIME": {"format": "yyyy-MM-dd HH:mm:ss", "code": "yyyy-mm-dd hh:mm:ss", "excel_format": "yyyy-mm-dd hh:mm:ss"},
    "DATETIME_FRENCH": {"format": "dd/MM/yyyy HH:mm:ss", "code": "dd/mm/yyyy hh:mm:ss", "excel_format": "dd/mm/yyyy hh:mm:ss"},
    "DATE_ONLY": {"format": "yyyy-MM-dd", "code": "yyyy-mm-dd", "excel_format": "yyyy-mm-dd"},
    "TIME_ONLY": {"format": "HH:mm:ss", "code": "hh:mm:ss", "excel_format": "hh:mm:ss"},
    "SHORT": {"format": "dd/MM/yy", "code": "dd/mm/yy", "excel_format": "dd/mm/yy"},
    "CUSTOM": {"format": "", "code": "", "excel_format": ""}
}

# Configuration du logging
LOG_FILE = f'excel_compiler_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Dictionnaires pour l'internationalisation
TRANSLATIONS = {
    "fr": {
        "app_title": "Compilateur Excel Professionnel",
        "file_selection": "Sélection des fichiers",
        "compilation_options": "Options de compilation",
        "advanced_options": "Options avancées",
        "help": "Aide",
        "about": "À propos",
        "preview": "Aperçu",
        "date_format": "Format de date",
        "languages": "Langues",
        "no_directory": "Aucun répertoire sélectionné",
        "choose_directory": "Choisir un répertoire",
        "select_all_files": "Sélectionner tous les fichiers",
        "files_selected": "{} fichier(s) sélectionné(s)",
        "date_time": "Date et heure : {}",
        "header_start_row": "Ligne de début des en-têtes :",
        "header_rows": "Nombre de lignes d'en-tête :",
        "repeat_headers": "Répéter les en-têtes pour chaque fichier",
        "merge_headers": "Fusionner les en-têtes multi-niveaux",
        "add_filename": "Ajouter les noms des fichiers sources",
        "output_filename": "Nom du fichier de sortie :",
        "enable_verification": "Activer la vérification préliminaire des fichiers",
        "start_compilation": "Lancer la compilation",
        "remove_duplicates": "Supprimer les doublons",
        "remove_empty_rows": "Supprimer les lignes entièrement vides",
        "sort_data": "Trier les données",
        "sort_column": "Colonne de tri (ex: A, B, C) :",
        "auto_width": "Ajuster automatiquement la largeur des colonnes",
        "freeze_headers": "Figer les en-têtes",
        "file_formats": "Formats de fichiers supportés",
        "excel_files": "Fichiers Excel (.xlsx, .xls)",
        "csv_files": "Fichiers CSV (.csv)",
        "preview_data": "Prévisualiser les données avant compilation",
        "refresh_preview": "Actualiser l'aperçu",
        "preview_limited": "Aperçu limité aux {} premières lignes",
        "date_format_options": "Options de format de date",
        "date_format_standard": "Standard (AAAA-MM-JJ)",
        "date_format_french": "Français (JJ/MM/AAAA)",
        "date_format_us": "Américain (MM/JJ/AAAA)",
        "date_format_datetime": "Date et heure (AAAA-MM-JJ HH:MM:SS)",
        "date_format_date_only": "Date uniquement (AAAA-MM-JJ)",
        "date_format_time_only": "Heure uniquement (HH:MM:SS)",
        "date_format_short": "Format court (JJ/MM/AA)",
        "date_format_custom": "Personnalisé :",
        "language": "Langue",
        "french": "Français",
        "english": "Anglais",
        "apply": "Appliquer",
        "cancel": "Annuler",
        "ok": "OK",
        "error": "Erreur",
        "success": "Succès",
        "warning": "Avertissement",
        "no_files_selected": "Aucun fichier sélectionné",
        "select_files_message": "Veuillez sélectionner au moins un fichier à compiler.",
        "compilation_in_progress": "Compilation en cours...",
        "compilation_complete": "Compilation terminée. Fichier enregistré : {}",
        "compilation_failed": "Échec: {}",
        "no_data": "Aucune donnée à écrire.",
        "no_data_message": "Aucune donnée valide n'a pu être compilée.",
        "file_open_error": "Le fichier de sortie est ouvert dans une autre application.",
        "file_open_error_message": "Le fichier de sortie est ouvert dans une autre application. Veuillez le fermer et réessayer.",
        "verification_title": "Rapport de vérification des fichiers",
        "verification_header": "Vérification de {} fichiers",
        "compilable": "Compilables: {} ({}%)",
        "not_compilable": "Non compilables: {} ({}%)",
        "non_compilable_files": "Fichiers non compilables",
        "compilable_files": "Fichiers compilables",
        "error_resolution_tips": "Ces fichiers ne peuvent pas être compilés pour les raisons indiquées. Voici quelques conseils pour résoudre les problèmes courants :",
        "open_file_tip": "Fichier ouvert: Fermez le fichier dans Excel et réessayez",
        "protected_file_tip": "Fichier protégé: Désactivez la protection dans Excel (Révision > Protéger la feuille)",
        "header_structure_tip": "Structure d'en-tête incompatible: Vérifiez que le nombre de lignes d'en-tête est correct",
        "encoding_error_tip": "Erreur d'encodage: Réenregistrez le fichier CSV avec l'encodage UTF-8",
        "ignore_non_compilable": "Ignorer les non compilables et compiler",
        "filename": "Nom du fichier",
        "detected_issue": "Problème détecté",
        "compilation_report": "Rapport de compilation",
        "compilation_result": "Résultat de la compilation",
        "generated_file": "Fichier généré:",
        "compilation_rate": "Taux de compilation: {}% ({}/{})",
        "compiled_files": "Fichiers compilés ({})",
        "not_compiled_files": "Fichiers non compilés ({})",
        "all_files_compiled": "Tous les fichiers ont été compilés avec succès !",
        "ip_warning_title": "Propriété Intellectuelle - Avertissement",
        "ip_warning_content": "Cette application <b>Compilateur Excel Professionnel</b> est la propriété intellectuelle exclusive de:",
        "developer_name": "GOUNOU N'GOBI Chabi Zimé",
        "developer_title": "Data Manager & Data Analyst",
        "copyright_notice": "Tous droits réservés. Cette application est protégée par les lois sur le droit d'auteur et les traités internationaux sur la propriété intellectuelle.",
        "warning_important": "AVERTISSEMENT IMPORTANT:",
        "unauthorized_reproduction": "Toute reproduction non autorisée, distribution ou modification de cette application est strictement interdite.",
        "license_terms": "L'utilisation de cette application est soumise aux termes de la licence accordée par l'auteur.",
        "contact_info": "Pour toute question concernant les droits d'utilisation ou pour signaler une violation de la propriété intellectuelle, veuillez contacter l'auteur à l'adresse : zimkada@gmail.com.",
        "accept_conditions": "J'accepte ces conditions",
        "quit_app": "Quitter l'application",
        "file_menu": "Fichier",
        "edit_menu": "Édition",
        "tools_menu": "Outils",
        "language_menu": "Langue",
        "help_menu": "Aide",
        "open_dir": "Ouvrir un répertoire...",
        "save_settings": "Enregistrer les paramètres",
        "load_settings": "Charger les paramètres",
        "exit": "Quitter",
        "select_all": "Sélectionner tout",
        "deselect_all": "Désélectionner tout",
        "invert_selection": "Inverser la sélection",
        "preview_tool": "Aperçu des données",
        "compile_tool": "Compiler les fichiers",
        "about_help": "À propos...",
        "settings_saved": "Paramètres sauvegardés avec succès",
        "settings_loaded": "Paramètres chargés avec succès"
    },
    "en": {
        "app_title": "Professional Excel Compiler",
        "file_selection": "File Selection",
        "compilation_options": "Compilation Options",
        "advanced_options": "Advanced Options",
        "help": "Help",
        "about": "About",
        "preview": "Preview",
        "date_format": "Date Format",
        "languages": "Languages",
        "no_directory": "No directory selected",
        "choose_directory": "Choose directory",
        "select_all_files": "Select all files",
        "files_selected": "{} file(s) selected",
        "date_time": "Date and time: {}",
        "header_start_row": "Header start row:",
        "header_rows": "Number of header rows:",
        "repeat_headers": "Repeat headers for each file",
        "merge_headers": "Merge multi-level headers",
        "add_filename": "Add source filenames",
        "output_filename": "Output filename:",
        "enable_verification": "Enable preliminary file verification",
        "start_compilation": "Start compilation",
        "remove_duplicates": "Remove duplicates",
        "remove_empty_rows": "Remove empty rows",
        "sort_data": "Sort data",
        "sort_column": "Sort column (e.g. A, B, C):",
        "auto_width": "Auto-adjust column width",
        "freeze_headers": "Freeze headers",
        "file_formats": "Supported file formats",
        "excel_files": "Excel files (.xlsx, .xls)",
        "csv_files": "CSV files (.csv)",
        "preview_data": "Preview data before compilation",
        "refresh_preview": "Refresh preview",
        "preview_limited": "Preview limited to first {} rows",
        "date_format_options": "Date format options",
        "date_format_standard": "Standard (YYYY-MM-DD)",
        "date_format_french": "French (DD/MM/YYYY)",
        "date_format_us": "US (MM/DD/YYYY)",
        "date_format_datetime": "Date and time (YYYY-MM-DD HH:MM:SS)",
        "date_format_date_only": "Date only (YYYY-MM-DD)",
        "date_format_time_only": "Time only (HH:MM:SS)",
        "date_format_short": "Short format (DD/MM/YY)",
        "date_format_custom": "Custom:",
        "language": "Language:",
        "french": "French",
        "english": "English",
        "apply": "Apply",
        "cancel": "Cancel",
        "ok": "OK",
        "error": "Error",
        "success": "Success",
        "warning": "Warning",
        "no_files_selected": "No files selected",
        "select_files_message": "Please select at least one file to compile.",
        "compilation_in_progress": "Compilation in progress...",
        "compilation_complete": "Compilation complete. File saved: {}",
        "compilation_failed": "Failed: {}",
        "no_data": "No data to write.",
        "no_data_message": "No valid data could be compiled.",
        "file_open_error": "The output file is open in another application.",
        "file_open_error_message": "The output file is open in another application. Please close it and try again.",
        "verification_title": "File Verification Report",
        "verification_header": "Verification of {} files",
        "compilable": "Compilable: {} ({}%)",
        "not_compilable": "Not compilable: {} ({}%)",
        "non_compilable_files": "Non-compilable files",
        "compilable_files": "Compilable files",
        "error_resolution_tips": "These files cannot be compiled for the reasons indicated. Here are some tips to resolve common problems:",
        "open_file_tip": "Open file: Close the file in Excel and try again",
        "protected_file_tip": "Protected file: Disable protection in Excel (Review > Protect Sheet)",
        "header_structure_tip": "Incompatible header structure: Check that the number of header rows is correct",
        "encoding_error_tip": "Encoding error: Resave the CSV file with UTF-8 encoding",
        "ignore_non_compilable": "Ignore non-compilable and compile",
        "filename": "Filename",
        "detected_issue": "Detected issue",
        "compilation_report": "Compilation Report",
        "compilation_result": "Compilation Result",
        "generated_file": "Generated file:",
        "compilation_rate": "Compilation rate: {}% ({}/{})",
        "compiled_files": "Compiled files ({})",
        "not_compiled_files": "Files not compiled ({})",
        "all_files_compiled": "All files were successfully compiled!",
        "ip_warning_title": "Intellectual Property - Warning",
        "ip_warning_content": "This <b>Professional Excel Compiler</b> application is the exclusive intellectual property of:",
        "developer_name": "GOUNOU N'GOBI Chabi Zimé",
        "developer_title": "Data Manager & Data Analyst",
        "copyright_notice": "All rights reserved. This application is protected by copyright laws and international intellectual property treaties.",
        "warning_important": "IMPORTANT WARNING:",
        "unauthorized_reproduction": "Any unauthorized reproduction, distribution or modification of this application is strictly prohibited.",
        "license_terms": "Use of this application is subject to the terms of the license granted by the author.",
        "contact_info": "For any questions regarding usage rights or to report a violation of intellectual property, please contact the author at: zimkada@gmail.com.",
        "accept_conditions": "I accept these conditions",
        "quit_app": "Quit application",
        "file_menu": "File",
        "edit_menu": "Edit",
        "tools_menu": "Tools",
        "language_menu": "Language",
        "help_menu": "Help",
        "open_dir": "Open directory...",
        "save_settings": "Save settings",
        "load_settings": "Load settings",
        "exit": "Exit",
        "select_all": "Select all",
        "deselect_all": "Deselect all",
        "invert_selection": "Invert selection",
        "preview_tool": "Data preview",
        "compile_tool": "Compile files",
        "about_help": "About...",
        "settings_saved": "Settings saved successfully",
        "settings_loaded": "Settings loaded successfully"
    }
}

class TranslationManager:
    """
    Gestionnaire de traduction pour l'application.
    Permet de changer la langue de l'interface utilisateur.
    """
    _instance = None
    
    def __new__(cls):
        """Implémentation du pattern Singleton."""
        if cls._instance is None:
            cls._instance = super(TranslationManager, cls).__new__(cls)
            cls._instance._initialize()
        return cls._instance
    
    def _initialize(self):
        """Initialise le gestionnaire de traduction."""
        self.current_language = "fr"  # Langue par défaut
        self.translations = TRANSLATIONS
        self.language_changed_callbacks = []
    
    def set_language(self, language_code):
        """
        Change la langue courante.
        
        Args:
            language_code: Code de la langue (fr, en)
        """
        if language_code in self.translations:
            self.current_language = language_code
            
            # Appeler tous les callbacks enregistrés
            for callback in self.language_changed_callbacks:
                callback()
    
    def get_text(self, key, *args):
        """
        Obtient le texte traduit pour une clé donnée.
        
        Args:
            key: Clé de traduction
            *args: Arguments de formatage optionnels
            
        Returns:
            Texte traduit
        """
        if key in self.translations[self.current_language]:
            text = self.translations[self.current_language][key]
            if args:
                return text.format(*args)
            return text
        return key
    
    def register_language_changed_callback(self, callback):
        """
        Enregistre un callback à appeler lorsque la langue change.
        
        Args:
            callback: Fonction à appeler
        """
        if callback not in self.language_changed_callbacks:
            self.language_changed_callbacks.append(callback)
    
    def unregister_language_changed_callback(self, callback):
        """
        Supprime un callback enregistré.
        
        Args:
            callback: Fonction à supprimer
        """
        if callback in self.language_changed_callbacks:
            self.language_changed_callbacks.remove(callback)

class SettingsManager:
    """
    Gestionnaire des paramètres de l'application.
    Permet de sauvegarder et charger les préférences utilisateur.
    """
    _instance = None
    
    def __new__(cls):
        """Implémentation du pattern Singleton."""
        if cls._instance is None:
            cls._instance = super(SettingsManager, cls).__new__(cls)
            cls._instance._initialize()
        return cls._instance
    
    def _initialize(self):
        """Initialise le gestionnaire de paramètres."""
        self.settings = QSettings("GounouNGobi", "ExcelCompiler")
    
    def save_settings(self, window):
        """
        Sauvegarde les paramètres actuels de l'application.
        
        Args:
            window: Fenêtre principale de l'application
        """
        settings = self.settings
        
        # Sauvegarde des options générales
        settings.setValue("language", TranslationManager().current_language)
        settings.setValue("headerStartRow", window.spinbox_header_start.value())
        settings.setValue("headerRows", window.spinbox_header.value())
        settings.setValue("repeatHeaders", window.checkbox_repeat_header.isChecked())
        settings.setValue("mergeHeaders", window.checkbox_merge_headers.isChecked())
        settings.setValue("addFilename", window.checkbox_add_filename.isChecked())
        settings.setValue("outputFilename", window.lineedit_output_name.text())
        settings.setValue("verifyFiles", window.checkbox_verify_files.isChecked())
        
        # Options avancées
        settings.setValue("removeDuplicates", window.checkbox_remove_duplicates.isChecked())
        settings.setValue("removeEmptyRows", window.checkbox_remove_empty_rows.isChecked())
        settings.setValue("sortData", window.checkbox_sort_data.isChecked())
        settings.setValue("sortColumn", window.lineedit_sort_column.text())
        settings.setValue("autoWidth", window.checkbox_auto_width.isChecked())
        settings.setValue("freezeHeader", window.checkbox_freeze_header.isChecked())
        settings.setValue("csvSupport", window.checkbox_csv.isChecked())
        
        # Option de format de date
        settings.setValue("dateFormat", window.date_format)
        
        # Sauvegarde du dernier répertoire utilisé
        if window.directory:
            settings.setValue("lastDirectory", window.directory)
        
        settings.sync()
    
    def load_settings(self, window):
        """
        Charge les paramètres sauvegardés.
        
        Args:
            window: Fenêtre principale de l'application
            
        Returns:
            bool: True si des paramètres ont été chargés, False sinon
        """
        settings = self.settings
        
        # Vérifier si des paramètres existent
        if not settings.contains("headerStartRow"):
            return False
        
        # Chargement de la langue
        language = settings.value("language", "fr")
        TranslationManager().set_language(language)
        
        # Chargement des options générales
        window.spinbox_header_start.setValue(int(settings.value("headerStartRow", 1)))
        window.spinbox_header.setValue(int(settings.value("headerRows", 1)))
        window.checkbox_repeat_header.setChecked(self._to_bool(settings.value("repeatHeaders", False)))
        window.checkbox_merge_headers.setChecked(self._to_bool(settings.value("mergeHeaders", False)))
        window.checkbox_add_filename.setChecked(self._to_bool(settings.value("addFilename", True)))
        window.lineedit_output_name.setText(settings.value("outputFilename", "compilation.xlsx"))
        window.checkbox_verify_files.setChecked(self._to_bool(settings.value("verifyFiles", True)))
        
        # Options avancées
        window.checkbox_remove_duplicates.setChecked(self._to_bool(settings.value("removeDuplicates", True)))
        window.checkbox_remove_empty_rows.setChecked(self._to_bool(settings.value("removeEmptyRows", False)))
        window.checkbox_sort_data.setChecked(self._to_bool(settings.value("sortData", False)))
        window.lineedit_sort_column.setText(settings.value("sortColumn", "A"))
        window.lineedit_sort_column.setEnabled(window.checkbox_sort_data.isChecked())
        window.checkbox_auto_width.setChecked(self._to_bool(settings.value("autoWidth", True)))
        window.checkbox_freeze_header.setChecked(self._to_bool(settings.value("freezeHeader", True)))
        window.checkbox_csv.setChecked(self._to_bool(settings.value("csvSupport", True)))
        
        # Option de format de date
        window.date_format = settings.value("dateFormat", "FRENCH")
    
        
        # Chargement du dernier répertoire utilisé
        last_directory = settings.value("lastDirectory", "")
        if last_directory and os.path.isdir(last_directory):
            window.directory = last_directory
            window.label_directory.setText(last_directory)
            window.load_files()
        
        return True
    
    def _to_bool(self, value):
        """
        Convertit une valeur en booléen.
        
        Args:
            value: Valeur à convertir
            
        Returns:
            bool: Valeur convertie
        """
        if isinstance(value, bool):
            return value
        return value.lower() in ("true", "1", "yes", "y", "t")
    
class FileVerification:
    """
    Classe utilitaire pour vérifier la compatibilité des fichiers Excel et CSV.
    """
    
    @staticmethod
    def verify_excel_file(file_path: str, header_start_row: int, header_rows: int) -> Tuple[bool, str]:
        """
        Vérifie si un fichier Excel est compatible pour la compilation.
        
        Args:
            file_path: Chemin complet du fichier à vérifier
            header_start_row: Ligne de début des en-têtes
            header_rows: Nombre de lignes d'en-tête
            
        Returns:
            Tuple[bool, str]: (est_compatible, message_d'erreur)
        """
        try:
            # Essai d'ouverture du fichier
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active
            
            # Vérification du nombre de lignes
            if ws.max_row < header_start_row + header_rows:
                return False, f"Structure d'en-tête incompatible: le fichier n'a que {ws.max_row} lignes"
                
            # Vérification de la protection du fichier
            if hasattr(ws, 'protection') and ws.protection.sheet:
                return False, "Le fichier est protégé en écriture"
                
            # Autres vérifications possibles...
            
            wb.close()
            return True, "Fichier compatible"
            
        except PermissionError:
            return False, "Le fichier est ouvert dans une autre application"
        except Exception as e:
            return False, f"Erreur lors de la vérification: {str(e)}"
    
    @staticmethod
    def verify_csv_file(file_path: str, header_start_row: int, header_rows: int) -> Tuple[bool, str]:
        """
        Vérifie si un fichier CSV est compatible pour la compilation.
        
        Args:
            file_path: Chemin complet du fichier à vérifier
            header_start_row: Ligne de début des en-têtes
            header_rows: Nombre de lignes d'en-tête
            
        Returns:
            Tuple[bool, str]: (est_compatible, message_d'erreur)
        """
        try:
            # Compter les lignes du fichier
            with open(file_path, 'r', newline='', encoding='utf-8-sig') as csvfile:
                reader = csv.reader(csvfile)
                row_count = sum(1 for _ in reader)
                
            if row_count < header_start_row + header_rows:
                return False, f"Structure d'en-tête incompatible: le fichier n'a que {row_count} lignes"
                
            return True, "Fichier compatible"
            
        except UnicodeDecodeError:
            # Essayer avec une autre encodage
            try:
                with open(file_path, 'r', newline='', encoding='latin-1') as csvfile:
                    reader = csv.reader(csvfile)
                    row_count = sum(1 for _ in reader)
                    
                if row_count < header_start_row + header_rows:
                    return False, f"Structure d'en-tête incompatible: le fichier n'a que {row_count} lignes"
                    
                return True, "Fichier compatible (encodage latin-1)"
            except Exception as e:
                return False, f"Erreur d'encodage: {str(e)}"
                
        except PermissionError:
            return False, "Le fichier est ouvert dans une autre application"
        except Exception as e:
            return False, f"Erreur lors de la vérification: {str(e)}"


class CompilationWorker(QThread):
    """
    Thread de travail qui gère la compilation des fichiers Excel et CSV.
    Permet de ne pas bloquer l'interface utilisateur pendant le traitement.
    """
    
    progress = pyqtSignal(int)  # Signal de progression
    error = pyqtSignal(str)     # Signal d'erreur
    finished = pyqtSignal(tuple)  # Signal de fin avec les données compilées
    
    def __init__(self, files, directory, header_start_row, header_rows, add_filename=True,
                 sort_data=False, sort_column=0, repeat_headers=False, remove_empty_rows=False,
                 remove_duplicates=False, date_format="FRENCH", parent=None):
        """
        Initialisation du worker de compilation.
        
        Args:
            files: Liste des noms de fichiers à compiler
            directory: Répertoire contenant les fichiers
            header_start_row: Ligne de début des en-têtes
            header_rows: Nombre de lignes d'en-tête
            add_filename: Ajouter le nom du fichier source comme colonne
            sort_data: Trier les données
            sort_column: Colonne de tri (index)
            repeat_headers: Répéter les en-têtes pour chaque fichier
            remove_empty_rows: Supprimer les lignes vides
            remove_duplicates: Supprimer les doublons
            date_format: Format de date à utiliser
            parent: Widget parent
        """
        super().__init__(parent)
        self.files = files
        self.directory = directory
        self.header_start_row = header_start_row
        self.header_rows = header_rows
        self.add_filename = add_filename
        self.sort_data = sort_data
        self.sort_column = sort_column
        self.repeat_headers = repeat_headers
        self.remove_empty_rows = remove_empty_rows
        self.remove_duplicates = remove_duplicates
        self.date_format = date_format
        
    def run(self):
        """
        Méthode principale exécutée dans le thread.
        Réalise la compilation des fichiers.
        """
        combined_data = []
        headers = None
        merged_cells = []
        preliminary_info = []
        
        successful_files = []
        failed_files = []
        
        for i, file in enumerate(self.files):
            try:
                file_path = os.path.join(self.directory, file)
                
                # Traitement différent selon le type de fichier
                if file.lower().endswith(('.xlsx', '.xls')):
                    result = self._process_excel_file(file_path, i, headers, preliminary_info)
                elif file.lower().endswith('.csv'):
                    result = self._process_csv_file(file_path, i, headers, preliminary_info)
                else:
                    raise ValueError(f"Format de fichier non pris en charge: {file}")
                    
                if result:
                    file_headers, file_data, file_merged_cells = result
                    
                    # Mise à jour des en-têtes globales si nécessaire
                    if headers is None:
                        headers = file_headers
                        merged_cells = file_merged_cells
                        
                    # Ajout des données au résultat combiné    
                    combined_data.extend(file_data)
                    successful_files.append(file)
                
                self.progress.emit(i + 1)
                
            except Exception as e:
                logging.error(f"Erreur lors du traitement du fichier {file}: {str(e)}")
                logging.error(traceback.format_exc())
                failed_files.append((file, str(e)))
                self.error.emit(f"Erreur avec le fichier {file}: {str(e)}")
                continue
        
        # Traitement post-compilation
        if combined_data and headers:
            # Suppression des doublons si demandé
            if self.remove_duplicates:
                combined_data = self._remove_duplicate_rows(combined_data)
                
            # Tri des données si l'option est activée
            if self.sort_data and self.sort_column < len(headers[-1]):
                combined_data = self._sort_data(combined_data, headers)
            
            # Ajout du nom du fichier source à l'en-tête si nécessaire    
            if self.add_filename and headers:
                headers[-1].append("Fichier source")
        
        # Émission du signal de fin avec toutes les données
        self.finished.emit((preliminary_info, headers, combined_data, merged_cells, successful_files, failed_files))
            
    def _process_excel_file(self, file_path, file_index, global_headers, preliminary_info):
        """
        Traite un fichier Excel et extrait ses données.
        
        Args:
            file_path: Chemin du fichier
            file_index: Index du fichier dans la liste
            global_headers: En-têtes déjà établies (pour les fichiers suivants)
            preliminary_info: Informations préliminaires à collecter du premier fichier
            
        Returns:
            Tuple contenant les en-têtes, données et cellules fusionnées du fichier
        """
        # Pour le premier fichier, on a besoin des merged_cells donc on ne met pas read_only
        if file_index == 0:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        else:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            
        ws = wb.active
        file_merged_cells = []
        file_headers = []
        
        # Capture des informations préliminaires du premier fichier
        if file_index == 0 and self.header_start_row > 1:
            for row in range(1, self.header_start_row):
                row_data = []
                for cell in ws[row]:
                    row_data.append(cell.value)
                preliminary_info.append(row_data)
        
        # Capture des en-têtes du premier fichier ou utilisation des en-têtes globales
        if global_headers is None:
            for row in range(self.header_start_row, self.header_start_row + self.header_rows):
                header_row = []
                for cell in ws[row]:
                    header_row.append(cell.value)
                file_headers.append(header_row)
                
                # Capture des merged_cells seulement pour le premier fichier
                if hasattr(ws, 'merged_cells'):
                    for merged_range in ws.merged_cells.ranges:
                        if merged_range.min_row <= (self.header_start_row + self.header_rows):
                            file_merged_cells.append(merged_range)
        else:
            file_headers = global_headers
            
        # Ajout des données du fichier
        file_data = self._extract_excel_data(ws, file_path.split(os.sep)[-1])
        
        # Fermer le workbook pour libérer la mémoire
        wb.close()
        
        return file_headers, file_data, file_merged_cells
        
    def _extract_excel_data(self, worksheet, filename):
        """
        Extrait les données d'une feuille Excel.
        
        Args:
            worksheet: Feuille de calcul à traiter
            filename: Nom du fichier source
            
        Returns:
            Liste des données extraites
        """
        data = []
        
        # Si répétition des en-têtes est activée et ce n'est pas le premier fichier
        if self.repeat_headers and len(data) > 0:
            # Ajouter une ligne vide comme séparateur
            data.append([None] * (len(worksheet[self.header_start_row]) + (1 if self.add_filename else 0)))
            
            # Ajouter les en-têtes
            for row in range(self.header_start_row, self.header_start_row + self.header_rows):
                row_data = [cell.value for cell in worksheet[row]]
                if self.add_filename:
                    row_data.append(None)  # Pas de nom de fichier dans l'en-tête répété
                data.append(row_data)
        
        # Ajout des données
        for row in worksheet.iter_rows(min_row=self.header_start_row + self.header_rows):
            row_data = [cell.value for cell in row]
            
            # Ajouter le nom du fichier si demandé
            if self.add_filename:
                row_data.append(filename)
            
            # Vérifier si la ligne n'est pas vide avant de l'ajouter
            if not self.remove_empty_rows or not all(
                cell is None or str(cell).strip() == "" 
                for cell in row_data[:-1 if self.add_filename else None]
            ):
                data.append(row_data)
                
        return data
        
    def _process_csv_file(self, file_path, file_index, global_headers, preliminary_info):
        """
        Traite un fichier CSV et extrait ses données.
        
        Args:
            file_path: Chemin du fichier
            file_index: Index du fichier dans la liste
            global_headers: En-têtes déjà établies (pour les fichiers suivants)
            preliminary_info: Informations préliminaires à collecter du premier fichier
            
        Returns:
            Tuple contenant les en-têtes et données du fichier
        """
        # Détection de l'encodage du fichier
        encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    # Détection du délimiteur
                    sample = f.read(4096)
                    sniffer = csv.Sniffer()
                    dialect = sniffer.sniff(sample)
                    delimiter = dialect.delimiter
                    break
            except Exception:
                continue
        else:
            raise ValueError(f"Impossible de déterminer l'encodage du fichier CSV: {file_path}")
        
        # Lecture du CSV avec pandas
        df = pd.read_csv(file_path, delimiter=delimiter, header=None, encoding=encoding)
        
        file_headers = []
        file_merged_cells = []  # Toujours vide pour les CSV
        
        # Capture des informations préliminaires du premier fichier
        if file_index == 0 and self.header_start_row > 1:
            for row in range(0, self.header_start_row - 1):
                if row < len(df):
                    preliminary_info.append(df.iloc[row].tolist())
        
        # Capture des en-têtes du premier fichier ou utilisation des en-têtes globales
        if global_headers is None:
            for row in range(self.header_start_row - 1, self.header_start_row - 1 + self.header_rows):
                if row < len(df):
                    file_headers.append(df.iloc[row].tolist())
        else:
            file_headers = global_headers
            
        # Extraction des données
        file_data = []
        
        # Si répétition des en-têtes est activée et ce n'est pas le premier fichier
        if self.repeat_headers and len(file_data) > 0:
            # Ajouter une ligne vide comme séparateur
            file_data.append([None] * (len(file_headers[-1]) + (1 if self.add_filename else 0)))
            
            # Ajouter les en-têtes
            for header_row in file_headers:
                row_data = header_row.copy()
                if self.add_filename:
                    row_data.append(None)
                file_data.append(row_data)
        
        # Ajout des données
        for row in range(self.header_start_row - 1 + self.header_rows, len(df)):
            row_data = df.iloc[row].tolist()
            
            # Ajouter le nom du fichier si demandé
            if self.add_filename:
                row_data.append(os.path.basename(file_path))
            
            # Vérifier si la ligne n'est pas vide avant de l'ajouter
            if not self.remove_empty_rows or not all(
                pd.isna(cell) or str(cell).strip() == "" 
                for cell in row_data[:-1 if self.add_filename else None]
            ):
                file_data.append(row_data)
                
        return file_headers, file_data, file_merged_cells
    
    def _remove_duplicate_rows(self, data):
        """
        Supprime les lignes en double dans les données.
        
        Args:
            data: Liste des données à filtrer
            
        Returns:
            Liste des données sans doublons
        """
        unique_data = []
        seen = set()
        
        for row in data:
            # Convertir la ligne en tuple pour pouvoir l'ajouter à un set
            row_tuple = tuple(str(cell) if cell is not None else None for cell in row)
            
            if row_tuple not in seen:
                seen.add(row_tuple)
                unique_data.append(row)
                
        return unique_data
        
    def _sort_data(self, data, headers):
        """
        Trie les données selon la colonne spécifiée.
        
        Args:
            data: Données à trier
            headers: En-têtes pour déterminer le nombre de colonnes
            
        Returns:
            Données triées
        """
        try:
            sort_idx = self.sort_column - 1
            
            if self.repeat_headers:
                # Cas spécial: préserver les sections avec en-têtes répétés
                sections = []
                current_section = []
                
                for row in data:
                    # Une ligne entièrement vide indique un séparateur de section
                    if all(cell is None for cell in row):
                        if current_section:
                            sections.append(current_section)
                        current_section = [row]  # Garder la ligne vide
                    else:
                        current_section.append(row)
                
                if current_section:
                    sections.append(current_section)
                
                # Trier chaque section individuellement
                for section in sections:
                    # Séparer l'en-tête et les données
                    header_rows = []
                    data_rows = []
                    
                    for i, row in enumerate(section):
                        # La première ligne est le séparateur, puis viennent les en-têtes
                        if i < self.header_rows + 1:
                            header_rows.append(row)
                        else:
                            data_rows.append(row)
                    
                    # Trier uniquement les données
                    data_rows.sort(key=lambda x: (x[sort_idx] is None, x[sort_idx]))
                    
                    # Recombiner
                    section.clear()
                    section.extend(header_rows)
                    section.extend(data_rows)
                
                # Aplatir les sections
                sorted_data = []
                for section in sections:
                    sorted_data.extend(section)
                
                return sorted_data
            else:
                # Cas simple: trier toutes les données
                return sorted(data, key=lambda x: (x[sort_idx] is None, x[sort_idx]))
                
        except Exception as e:
            logging.warning(f"Erreur lors du tri : {str(e)}")
            return data


class ExcelFormatter:
    """
    Classe responsable du formatage du fichier Excel de sortie.
    """
    
    # Styles prédéfinis
    HEADER_STYLE = {
        'font': Font(bold=True, color=EXCEL_COLORS["LIGHT_TEXT"]),
        'fill': PatternFill(start_color=EXCEL_COLORS["PRIMARY"], end_color=EXCEL_COLORS["PRIMARY"], fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center')
    }
    
    PRELIMINARY_STYLE = {
        'font': Font(italic=True)
    }
    
    DATA_BORDER = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    
    @staticmethod
    def write_preliminary_info(worksheet, preliminary_info):
        """
        Écrit les informations préliminaires dans la feuille de calcul.
        
        Args:
            worksheet: Feuille de calcul à modifier
            preliminary_info: Données préliminaires à écrire
            
        Returns:
            Ligne courante après écriture
        """
        current_row = 1
        
        for row_data in preliminary_info:
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                cell.font = ExcelFormatter.PRELIMINARY_STYLE['font']
            current_row += 1
            
        return current_row
    
    @staticmethod
    def write_headers(worksheet, headers, start_row):
        """
        Écrit les en-têtes dans la feuille de calcul avec le style approprié.
        
        Args:
            worksheet: Feuille de calcul à modifier
            headers: Données d'en-tête à écrire
            start_row: Ligne de début pour écrire les en-têtes
            
        Returns:
            Ligne courante après écriture
        """
        current_row = start_row
        
        for header_row in headers:
            for col_idx, value in enumerate(header_row, 1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                cell.font = ExcelFormatter.HEADER_STYLE['font']
                cell.fill = ExcelFormatter.HEADER_STYLE['fill']
                cell.alignment = ExcelFormatter.HEADER_STYLE['alignment']
            current_row += 1
            
        return current_row
    
    @staticmethod
    def write_data(worksheet, data, start_row, date_format="FRENCH"):
        """
        Écrit les données dans la feuille de calcul avec le style approprié.
        
        Args:
            worksheet: Feuille de calcul à modifier
            data: Données à écrire
            start_row: Ligne de début pour écrire les données
            date_format: Format de date à utiliser (par défaut FRENCH)
            
        Returns:
            Ligne courante après écriture
        """
        current_row = start_row
        
        excel_date_format = DATE_FORMATS.get(date_format, DATE_FORMATS["FRENCH"])["excel_format"]
        
        for row_data in data:
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                cell.border = ExcelFormatter.DATA_BORDER
                
                # Appliquer le format de date si nécessaire
                if isinstance(value, datetime):
                    cell.number_format = excel_date_format
                    
            current_row += 1
            
        return current_row
    
    @staticmethod
    def apply_merged_cells(worksheet, merged_ranges, header_start_original, header_start_new):
        """
        Applique les fusions de cellules en ajustant les numéros de ligne.
        
        Args:
            worksheet: Feuille de calcul à modifier
            merged_ranges: Plages de cellules à fusionner
            header_start_original: Ligne de début originale des en-têtes
            header_start_new: Nouvelle ligne de début des en-têtes
        """
        for merged_range in merged_ranges:
            # Ajuster les numéros de ligne
            adjusted_range = openpyxl.worksheet.cell_range.CellRange(
                min_col=merged_range.min_col,
                min_row=merged_range.min_row - header_start_original + header_start_new,
                max_col=merged_range.max_col,
                max_row=merged_range.max_row - header_start_original + header_start_new
            )
            worksheet.merge_cells(range_string=adjusted_range.coord)
    
    @staticmethod
    def adjust_column_widths(worksheet):
        """
        Ajuste la largeur des colonnes en fonction du contenu.
        
        Args:
            worksheet: Feuille de calcul à modifier
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                    
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
    
    @staticmethod
    def freeze_header(worksheet, freeze_row):
        """
        Fige les volets à la ligne spécifiée.
        
        Args:
            worksheet: Feuille de calcul à modifier
            freeze_row: Ligne à partir de laquelle figer les volets
        """
        worksheet.freeze_panes = worksheet.cell(row=freeze_row, column=1)

class PreviewDialog(QDialog):
    """
    Boîte de dialogue pour prévisualiser les données avant compilation.
    """
    
    def __init__(self, parent, directory, files, header_start_row, header_rows):
        """
        Initialise la boîte de dialogue de prévisualisation.
        
        Args:
            parent: Widget parent
            directory: Répertoire contenant les fichiers
            files: Liste des fichiers à prévisualiser
            header_start_row: Ligne de début des en-têtes
            header_rows: Nombre de lignes d'en-tête
        """
        super().__init__(parent)
        self.translate = TranslationManager().get_text
        self.setWindowTitle(self.translate("preview"))
        self.resize(1000, 700)
        
        self.directory = directory
        self.files = files
        self.header_start_row = header_start_row
        self.header_rows = header_rows
        
        self.selected_file = None
        self.preview_data = None
        self.max_preview_rows = 200  # Limite de lignes pour la prévisualisation
        
        self.init_ui()
        
        # Charger le premier fichier s'il existe
        if files:
            self.file_combo.setCurrentIndex(0)
            self.selected_file = files[0]
            self.load_preview()
    
    def init_ui(self):
        """Initialise l'interface utilisateur de la boîte de dialogue."""
        layout = QVBoxLayout()
        
        # Sélection du fichier à prévisualiser
        file_layout = QHBoxLayout()
        file_label = QLabel(self.translate("filename") + ":")
        self.file_combo = QComboBox()
        self.file_combo.addItems(self.files)
        self.file_combo.currentIndexChanged.connect(self.on_file_changed)
        
        self.refresh_button = QPushButton(self.translate("refresh_preview"))
        self.refresh_button.clicked.connect(self.load_preview)
        
        file_layout.addWidget(file_label)
        file_layout.addWidget(self.file_combo, 1)
        file_layout.addWidget(self.refresh_button)
        
        layout.addLayout(file_layout)
        
        # Tableau de prévisualisation
        self.preview_table = QTableWidget()
        self.preview_table.setAlternatingRowColors(True)
        layout.addWidget(self.preview_table)
        
        # Informations sur la prévisualisation
        self.info_label = QLabel(self.translate("preview_limited", self.max_preview_rows))
        layout.addWidget(self.info_label)
        
        # Bouton de fermeture
        button_layout = QHBoxLayout()
        close_button = QPushButton(self.translate("ok"))
        close_button.clicked.connect(self.accept)
        button_layout.addStretch()
        button_layout.addWidget(close_button)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def on_file_changed(self, index):
        """
        Appelé lorsque l'utilisateur change de fichier dans le combobox.
        
        Args:
            index: Index du fichier sélectionné
        """
        if index >= 0 and index < len(self.files):
            self.selected_file = self.files[index]
            self.load_preview()
    
    def load_preview(self):
        """Charge et affiche la prévisualisation du fichier sélectionné."""
        if not self.selected_file:
            return
        
        file_path = os.path.join(self.directory, self.selected_file)
        
        try:
            if self.selected_file.lower().endswith(('.xlsx', '.xls')):
                self.load_excel_preview(file_path)
            elif self.selected_file.lower().endswith('.csv'):
                self.load_csv_preview(file_path)
        except Exception as e:
            QMessageBox.warning(
                self,
                self.translate("error"),
                f"{self.translate('error')}: {str(e)}"
            )
    
    def load_excel_preview(self, file_path):
        """
        Charge la prévisualisation d'un fichier Excel.
        
        Args:
            file_path: Chemin du fichier Excel à prévisualiser
        """
        try:
            # Charger le fichier Excel
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active
            
            # Obtenir les données (limité au nombre max de lignes)
            preview_data = []
            headers = []
            
            # Récupérer les en-têtes
            for row in range(self.header_start_row, self.header_start_row + self.header_rows):
                header_row = []
                for cell in ws[row]:
                    header_row.append(cell.value)
                headers.append(header_row)
            
            # Récupérer les données
            row_count = 0
            for row in ws.iter_rows(min_row=self.header_start_row + self.header_rows):
                if row_count >= self.max_preview_rows:
                    break
                
                row_data = [cell.value for cell in row]
                preview_data.append(row_data)
                row_count += 1
            
            # Afficher les données dans le tableau
            self.display_preview(headers, preview_data)
            
            wb.close()
        except Exception as e:
            raise ValueError(f"Erreur lors de la lecture du fichier Excel: {str(e)}")
    
    def load_csv_preview(self, file_path):
        """
        Charge la prévisualisation d'un fichier CSV.
        
        Args:
            file_path: Chemin du fichier CSV à prévisualiser
        """
        try:
            # Détection de l'encodage du fichier
            encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
            delimiter = ','
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        sample = f.read(4096)
                        sniffer = csv.Sniffer()
                        delimiter = sniffer.sniff(sample).delimiter
                        break
                except Exception:
                    continue
            
            # Lecture du CSV avec pandas
            df = pd.read_csv(file_path, delimiter=delimiter, header=None, encoding=encoding)
            
            # Limiter le nombre de lignes
            if len(df) > self.max_preview_rows + self.header_start_row + self.header_rows:
                df = df.iloc[:(self.max_preview_rows + self.header_start_row + self.header_rows)]
            
            # Récupérer les en-têtes
            headers = []
            for row in range(self.header_start_row - 1, self.header_start_row - 1 + self.header_rows):
                if row < len(df):
                    headers.append(df.iloc[row].tolist())
            
            # Récupérer les données
            preview_data = []
            for row in range(self.header_start_row - 1 + self.header_rows, len(df)):
                preview_data.append(df.iloc[row].tolist())
            
            # Afficher les données dans le tableau
            self.display_preview(headers, preview_data)
            
        except Exception as e:
            raise ValueError(f"Erreur lors de la lecture du fichier CSV: {str(e)}")
    
    def display_preview(self, headers, data):
        """
        Affiche les données dans le tableau de prévisualisation.
        
        Args:
            headers: Données d'en-tête
            data: Données à afficher
        """
        if not headers or not headers[0]:
            return
        
        # Configurer le tableau
        self.preview_table.clear()
        self.preview_table.setRowCount(len(data))
        self.preview_table.setColumnCount(len(headers[-1]))
        
        # Définir les en-têtes des colonnes
        self.preview_table.setHorizontalHeaderLabels([str(h) if h is not None else "" for h in headers[-1]])
        
        # Ajouter les données
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data[:len(headers[-1])]):  # Limiter aux colonnes d'en-tête
                # Créer un élément de tableau avec la valeur formatée
                item = QTableWidgetItem(str(value) if value is not None else "")
                self.preview_table.setItem(row_idx, col_idx, item)
        
        # Ajuster la taille des colonnes
        self.preview_table.resizeColumnsToContents()
        
        # Mettre à jour l'étiquette d'information
        if len(data) >= self.max_preview_rows:
            self.info_label.setText(self.translate("preview_limited", self.max_preview_rows))
        else:
            self.info_label.setText(f"{len(data)} lignes affichées")


class DateFormatDialog(QDialog):
    """
    Boîte de dialogue pour choisir le format de date.
    """
    
    def __init__(self, parent, current_format="FRENCH"):
        """
        Initialise la boîte de dialogue de format de date.
        
        Args:
            parent: Widget parent
            current_format: Format de date actuellement sélectionné
        """
        super().__init__(parent)
        self.translate = TranslationManager().get_text
        self.setWindowTitle(self.translate("date_format"))
        self.resize(500, 400)
        
        self.current_format = current_format
        self.custom_format = ""
        self.selected_format = current_format
        
        self.init_ui()
    
    def init_ui(self):
        """Initialise l'interface utilisateur de la boîte de dialogue."""
        layout = QVBoxLayout()
        
        # Groupe de formats prédéfinis
        group = QGroupBox(self.translate("date_format_options"))
        group_layout = QVBoxLayout()
        
        # Créer les boutons radio pour chaque format prédéfini
        self.radio_group = QButtonGroup(self)
        formats = [
            ("STANDARD", "date_format_standard"),
            ("FRENCH", "date_format_french"),
            ("US", "date_format_us"),
            ("DATETIME", "date_format_datetime"),
            ("DATETIME_FRENCH", "date_format_datetime_french"),
            ("DATE_ONLY", "date_format_date_only"),
            ("TIME_ONLY", "date_format_time_only"),
            ("SHORT", "date_format_short"),
            ("CUSTOM", "date_format_custom")
        ]
        
        self.radio_buttons = {}
        
        for i, (format_key, label_key) in enumerate(formats):
            radio = QRadioButton(self.translate(label_key))
            self.radio_group.addButton(radio, i)
            self.radio_buttons[format_key] = radio
            
            if format_key == "CUSTOM":
                custom_layout = QHBoxLayout()
                custom_layout.addWidget(radio)
                self.custom_edit = QLineEdit()
                self.custom_edit.setPlaceholderText("dd/MM/yyyy HH:mm:ss")
                self.custom_edit.setEnabled(False)
                custom_layout.addWidget(self.custom_edit)
                group_layout.addLayout(custom_layout)
            else:
                group_layout.addWidget(radio)
        
        group.setLayout(group_layout)
        layout.addWidget(group)
        
        # Exemple avec la date actuelle
        example_layout = QHBoxLayout()
        example_layout.addWidget(QLabel(self.translate("preview") + ":"))
        self.example_label = QLabel()
        self.update_example()
        example_layout.addWidget(self.example_label)
        layout.addLayout(example_layout)
        
        # Boutons OK/Annuler
        button_layout = QHBoxLayout()
        ok_button = QPushButton(self.translate("ok"))
        ok_button.clicked.connect(self.accept)
        cancel_button = QPushButton(self.translate("cancel"))
        cancel_button.clicked.connect(self.reject)
        
        button_layout.addStretch()
        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        # Connecter les signaux
        self.radio_group.buttonClicked.connect(self.on_format_changed)
        self.custom_edit.textChanged.connect(self.on_custom_format_changed)
        
        # Sélectionner le format actuel
        if self.current_format in self.radio_buttons:
            self.radio_buttons[self.current_format].setChecked(True)
            if self.current_format == "CUSTOM":
                self.custom_edit.setEnabled(True)
    
    def on_format_changed(self, button):
        """
        Appelé lorsque l'utilisateur change de format.
        
        Args:
            button: Bouton radio sélectionné
        """
        for format_key, radio in self.radio_buttons.items():
            if radio == button:
                self.selected_format = format_key
                if format_key == "CUSTOM":
                    self.custom_edit.setEnabled(True)
                else:
                    self.custom_edit.setEnabled(False)
                break
        
        self.update_example()
    
    def on_custom_format_changed(self, text):
        """
        Appelé lorsque l'utilisateur modifie le format personnalisé.
        
        Args:
            text: Nouveau texte du format personnalisé
        """
        self.custom_format = text
        self.update_example()
    
    def update_example(self):
        """Met à jour l'exemple de format de date."""
        now = datetime.now()
        
        if self.selected_format == "CUSTOM":
            try:
                # Convertir le format de l'utilisateur en format de date Python
                user_format = self.custom_format
                # Remplacer les tokens de format
                py_format = user_format.replace("dd", "%d").replace("MM", "%m").replace("yyyy", "%Y")
                py_format = py_format.replace("HH", "%H").replace("mm", "%M").replace("ss", "%S")
                py_format = py_format.replace("yy", "%y")
                
                formatted_date = now.strftime(py_format)
                self.example_label.setText(formatted_date)
            except Exception:
                self.example_label.setText("Format invalide")
        else:
            # Utiliser le format prédéfini
            date_format = DATE_FORMATS[self.selected_format]["format"]
            
            # Convertir en format Python
            py_format = date_format.replace("dd", "%d").replace("MM", "%m").replace("yyyy", "%Y")
            py_format = py_format.replace("HH", "%H").replace("mm", "%M").replace("ss", "%S")
            py_format = py_format.replace("yy", "%y")
            
            formatted_date = now.strftime(py_format)
            self.example_label.setText(formatted_date)
    
    def get_selected_format(self):
        """
        Obtient le format de date sélectionné.
        
        Returns:
            str: Clé du format sélectionné
        """
        if self.selected_format == "CUSTOM":
            DATE_FORMATS["CUSTOM"]["format"] = self.custom_format
            
            # Générer un format Excel personnalisé
            excel_format = self.custom_format
            excel_format = excel_format.replace("dd", "dd").replace("MM", "mm").replace("yyyy", "yyyy")
            excel_format = excel_format.replace("HH", "hh").replace("mm", "mm").replace("ss", "ss")
            excel_format = excel_format.replace("yy", "yy")
            
            DATE_FORMATS["CUSTOM"]["excel_format"] = excel_format
            
        return self.selected_format


class LanguageDialog(QDialog):
    """
    Boîte de dialogue pour choisir la langue de l'interface.
    """
    
    def __init__(self, parent):
        """
        Initialise la boîte de dialogue de choix de langue.
        
        Args:
            parent: Widget parent
        """
        super().__init__(parent)
        self.translate = TranslationManager().get_text
        self.setWindowTitle(self.translate("languages"))
        self.resize(300, 200)
        
        self.selected_language = TranslationManager().current_language
        
        self.init_ui()
    
    def init_ui(self):
        """Initialise l'interface utilisateur de la boîte de dialogue."""
        layout = QVBoxLayout()
        
        # Groupe des langues
        group = QGroupBox(self.translate("language"))
        group_layout = QVBoxLayout()
        
        # Créer les boutons radio pour chaque langue
        self.radio_group = QButtonGroup(self)
        
        # Français
        self.radio_fr = QRadioButton(self.translate("french"))
        self.radio_group.addButton(self.radio_fr, 0)
        group_layout.addWidget(self.radio_fr)
        
        # Anglais
        self.radio_en = QRadioButton(self.translate("english"))
        self.radio_group.addButton(self.radio_en, 1)
        group_layout.addWidget(self.radio_en)
        
        group.setLayout(group_layout)
        layout.addWidget(group)
        
        # Boutons OK/Annuler
        button_layout = QHBoxLayout()
        apply_button = QPushButton(self.translate("apply"))
        apply_button.clicked.connect(self.accept)
        cancel_button = QPushButton(self.translate("cancel"))
        cancel_button.clicked.connect(self.reject)
        
        button_layout.addStretch()
        button_layout.addWidget(apply_button)
        button_layout.addWidget(cancel_button)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        # Sélectionner la langue actuelle
        if self.selected_language == "fr":
            self.radio_fr.setChecked(True)
        elif self.selected_language == "en":
            self.radio_en.setChecked(True)
        
        # Connecter les signaux
        self.radio_group.buttonClicked.connect(self.on_language_changed)
    
    def on_language_changed(self, button):
        """
        Appelé lorsque l'utilisateur change de langue.
        
        Args:
            button: Bouton radio sélectionné
        """
        if button == self.radio_fr:
            self.selected_language = "fr"
        elif button == self.radio_en:
            self.selected_language = "en"
    
    def get_selected_language(self):
        """
        Obtient la langue sélectionnée.
        
        Returns:
            str: Code de la langue sélectionnée
        """
        return self.selected_language


class VerificationReportDialog(QDialog):
    """
    Boîte de dialogue affichant le rapport de vérification des fichiers avant compilation.
    """
    
    def __init__(self, parent, compatible_files, incompatible_files):
        """
        Initialise la boîte de dialogue avec les résultats de la vérification.
        
        Args:
            parent: Widget parent
            compatible_files: Liste des fichiers compatibles
            incompatible_files: Liste des fichiers incompatibles avec raisons
        """
        super().__init__(parent)
        self.translate = TranslationManager().get_text
        self.setWindowTitle(self.translate("verification_title"))
        self.setMinimumSize(800, 600)
        self.compatible_files = compatible_files
        self.incompatible_files = incompatible_files
        self.continue_with_compatible = False
        
        self.init_ui()
    
    def init_ui(self):
        """Initialise l'interface utilisateur de la boîte de dialogue."""
        layout = QVBoxLayout()
        
        # En-tête avec statistiques
        total_files = len(self.compatible_files) + len(self.incompatible_files)
        header_label = QLabel(f"<h2>{self.translate('verification_header', total_files)}</h2>")
        header_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(header_label)
        
        compat_percent = len(self.compatible_files) * 100 / total_files if total_files > 0 else 0
        incompat_percent = len(self.incompatible_files) * 100 / total_files if total_files > 0 else 0
        
        stats_label = QLabel(
            f"<div style='text-align:center; margin:10px 0;'>"
            f"<span style='color:#{COLORS['SUCCESS']}; font-weight:bold;'>{self.translate('compilable', len(self.compatible_files), compat_percent)}</span> | "
            f"<span style='color:#{COLORS['WARNING']}; font-weight:bold;'>{self.translate('not_compilable', len(self.incompatible_files), incompat_percent)}</span>"
            f"</div>"
        )
        stats_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(stats_label)
        
        # Tableau des fichiers incompatibles
        if self.incompatible_files:
            group_incompatible = QGroupBox(self.translate("non_compilable_files"))
            group_layout = QVBoxLayout()
            
            # Instructions pour résoudre les problèmes
            help_label = QLabel(
                f"{self.translate('error_resolution_tips')}"
                "<ul>"
                f"<li><b>{self.translate('open_file_tip')}</b></li>"
                f"<li><b>{self.translate('protected_file_tip')}</b></li>"
                f"<li><b>{self.translate('header_structure_tip')}</b></li>"
                f"<li><b>{self.translate('encoding_error_tip')}</b></li>"
                "</ul>"
            )
            help_label.setWordWrap(True)
            group_layout.addWidget(help_label)
            
            table = QTableWidget()
            table.setColumnCount(2)
            table.setHorizontalHeaderLabels([self.translate("filename"), self.translate("detected_issue")])
            table.setRowCount(len(self.incompatible_files))
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
            
            for row, (file_name, reason) in enumerate(self.incompatible_files):
                table.setItem(row, 0, QTableWidgetItem(file_name))
                table.setItem(row, 1, QTableWidgetItem(reason))
                # Colorer la ligne en rouge pâle pour mettre en évidence
                for col in range(2):
                    table.item(row, col).setBackground(QBrush(QColor("#ffebee")))
            
            group_layout.addWidget(table)
            group_incompatible.setLayout(group_layout)
            layout.addWidget(group_incompatible)
        
        # Liste des fichiers compatibles
        if self.compatible_files:
            group_compatible = QGroupBox(self.translate("compilable_files"))
            group_layout = QVBoxLayout()
        
            list_widget = QListWidget()
            for file in self.compatible_files:
                item = QListWidgetItem(file)
                item.setBackground(QBrush(QColor("#e8f5e9")))  # Vert pâle
                list_widget.addItem(item)
            
            group_layout.addWidget(list_widget)
            group_compatible.setLayout(group_layout)
            layout.addWidget(group_compatible)
        
        # Boutons
        button_layout = QHBoxLayout()

        if self.incompatible_files and self.compatible_files:
            ignore_button = QPushButton(self.translate("ignore_non_compilable"))
            ignore_button.setStyleSheet(f"background-color: #{COLORS['INFO']}; color: white;")
            ignore_button.clicked.connect(self.continue_with_compatible_only)
            button_layout.addWidget(ignore_button)

        cancel_button = QPushButton(self.translate("cancel"))
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(cancel_button)

        layout.addLayout(button_layout)
        self.setLayout(layout)
    
    def continue_with_compatible_only(self):
        """
        Méthode appelée quand l'utilisateur choisit d'ignorer les fichiers incompatibles.
        """
        self.continue_with_compatible = True
        self.accept()

class CompilationReportDialog(QDialog):
    """
    Boîte de dialogue affichant le rapport après compilation.
    """
    
    def __init__(self, parent, successful_files, failed_files, output_path):
        """
        Initialise la boîte de dialogue avec les résultats de la compilation.
        
        Args:
            parent: Widget parent
            successful_files: Liste des fichiers compilés avec succès
            failed_files: Liste des fichiers échoués avec raisons
            output_path: Chemin du fichier de sortie
        """
        super().__init__(parent)
        self.translate = TranslationManager().get_text
        self.setWindowTitle(self.translate("compilation_report"))
        self.setMinimumSize(800, 600)
        self.successful_files = successful_files
        self.failed_files = failed_files
        self.output_path = output_path
        
        self.init_ui()
    
    def init_ui(self):
        """Initialise l'interface utilisateur de la boîte de dialogue."""
        layout = QVBoxLayout()
        
        # En-tête avec statistiques
        total_files = len(self.successful_files) + len(self.failed_files)
        header_label = QLabel(f"<h2>{self.translate('compilation_result')}</h2>")
        header_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(header_label)
        
        # Affichage du chemin du fichier de sortie
        output_label = QLabel(f"<div style='text-align:center; margin:5px 0;'>"
                             f"<b>{self.translate('generated_file')}</b> {self.output_path}"
                             f"</div>")
        output_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(output_label)
        
        # Statistiques
        percentage = len(self.successful_files) * 100 / total_files if total_files > 0 else 0
        stats_label = QLabel(f"<div style='text-align:center; margin:10px 0; font-size:16px;'>"
                            f"<b>{self.translate('compilation_rate', percentage, len(self.successful_files), total_files)}</b>"
                            f"</div>")
        stats_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(stats_label)
        
        # Icône de succès si tout a été compilé
        if percentage == 100:
            success_icon = QLabel()
            pixmap = self.style().standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton).pixmap(64, 64)
            success_icon.setPixmap(pixmap)
            success_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(success_icon)
            
            success_message = QLabel(f"<div style='text-align:center; color:green; font-weight:bold;'>"
                                    f"{self.translate('all_files_compiled')}"
                                    f"</div>")
            success_message.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(success_message)
        
        # Liste des fichiers compilés
        if self.successful_files:
            success_group = QGroupBox(self.translate("compiled_files", len(self.successful_files)))
            success_layout = QVBoxLayout()
            
            list_widget = QListWidget()
            for file in self.successful_files:
                item = QListWidgetItem(file)
                item.setBackground(QBrush(QColor("#e8f5e9")))  # Vert pâle
                list_widget.addItem(item)
            
            success_layout.addWidget(list_widget)
            success_group.setLayout(success_layout)
            layout.addWidget(success_group)
        
        # Tableau des échecs si existants
        if self.failed_files:
            failed_group = QGroupBox(self.translate("not_compiled_files", len(self.failed_files)))
            failed_layout = QVBoxLayout()
            
            table = QTableWidget()
            table.setColumnCount(2)
            table.setHorizontalHeaderLabels([self.translate("filename"), self.translate("error")])
            table.setRowCount(len(self.failed_files))
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
            
            for row, (file_name, reason) in enumerate(self.failed_files):
                table.setItem(row, 0, QTableWidgetItem(file_name))
                table.setItem(row, 1, QTableWidgetItem(reason))
                # Colorer la ligne en rouge pâle
                for col in range(2):
                    table.item(row, col).setBackground(QBrush(QColor("#ffebee")))
            
            failed_layout.addWidget(table)
            failed_group.setLayout(failed_layout)
            layout.addWidget(failed_group)
        
        # Bouton OK
        button = QPushButton(self.translate("ok"))
        button.clicked.connect(self.accept)
        layout.addWidget(button, alignment=Qt.AlignmentFlag.AlignCenter)
        
        self.setLayout(layout)


class IPWarningDialog(QDialog):
    """
    Boîte de dialogue d'avertissement concernant la propriété intellectuelle.
    """
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.translate = TranslationManager().get_text
        self.setWindowTitle(self.translate("ip_warning_title"))
        self.setMinimumSize(800, 600)
        
        self.init_ui()
    
    def init_ui(self):
        """Initialise l'interface utilisateur de la boîte de dialogue."""
        layout = QVBoxLayout()
        
        # Titre
        title_label = QLabel("<h1>" + self.translate("warning") + "</h1>")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # Contenu
        content = QLabel(
            "<p style='font-size: 14px; line-height: 1.5;'>"
            f"{self.translate('ip_warning_content')}<br><br>"
            "<div style='text-align: center; font-weight: bold; font-size: 16px; margin: 20px 0;'>"
            f"{self.translate('developer_name')}<br>"
            f"{self.translate('developer_title')}<br><br>"
            "</div>"
            f"{self.translate('copyright_notice')}<br><br>"
            
            f"<span style='color: #{COLORS['WARNING']}; font-weight: bold;'>{self.translate('warning_important')}</span><br>"
            "<ul>"
            f"<li>{self.translate('unauthorized_reproduction')}</li>"
            f"<li>{self.translate('license_terms')}</li>"
            "</ul><br>"
            
            f"{self.translate('contact_info')}"
            
            "</p>"
        )
        content.setWordWrap(True)
        content.setTextFormat(Qt.TextFormat.RichText)
        
        scroll = QScrollArea()
        scroll.setWidget(content)
        scroll.setWidgetResizable(True)
        layout.addWidget(scroll)
        
        # Boutons
        button_layout = QHBoxLayout()
        
        accept_button = QPushButton(self.translate("accept_conditions"))
        accept_button.clicked.connect(self.accept)
        accept_button.setDefault(True)
        
        exit_button = QPushButton(self.translate("quit_app"))
        exit_button.clicked.connect(self.reject)
        
        button_layout.addWidget(accept_button)
        button_layout.addWidget(exit_button)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)


class ModernExcelCompilerApp(QMainWindow):
    """
    Classe principale de l'application de compilation Excel.
    """
    def __init__(self):
        """Initialise l'application."""
        super().__init__()
        
        # Initialiser le gestionnaire de traduction
        self.translate = TranslationManager().get_text
        TranslationManager().register_language_changed_callback(self.update_ui_language)
        
        # Initialiser les variables
        self.setup_variables()
        
        # Configurer l'interface utilisateur
        self.setup_ui()
        self.create_menu()
        self.create_toolbar()
        self.create_main_layout()
        self.connect_signals()
        
        # Charger les paramètres sauvegardés
        SettingsManager().load_settings(self)
        
        logging.info("Application démarrée")
        
        # Afficher l'avertissement de propriété intellectuelle au démarrage
        self.show_ip_warning()
    
    def setup_variables(self):
        """Initialise les variables de l'application."""
        self.directory = ""
        self.files = []
        self.compilation_worker = None
        self.verification_enabled = True  # Par défaut, la vérification préliminaire est activée
        self.date_format = "FRENCH"  # Format de date par défaut


    def setup_ui(self):
        """Configure l'interface utilisateur principale."""
        self.setWindowTitle(self.translate("app_title"))
        self.setGeometry(50, 50, 1200, 700)
        self.setWindowIcon(QIcon(resource_path("icon.ico")))
        self.setMinimumSize(800, 600)
        self.apply_stylesheet()
    
    def apply_stylesheet(self):
        """Applique le style CSS à l'application."""
        self.setStyleSheet(f"""
            QMainWindow {{
                background-color: #{COLORS["BACKGROUND"]};
            }}
            
            /* Style pour la barre de menus */
            QMenuBar {{
                background-color: #{COLORS["PRIMARY"]};
                color: white;
                border: none;
            }}
            
            QMenuBar::item {{
                background-color: #{COLORS["PRIMARY"]};
                color: white;
            }}
            
            QMenuBar::item:selected {{
                background-color: #{COLORS["PRIMARY_DARK"]};
            }}
            
            QMenu {{
                background-color: white;
                border: 1px solid #{COLORS["PRIMARY"]};
            }}
            
            QMenu::item:selected {{
                background-color: #{COLORS["PRIMARY_LIGHT"]};
                color: white;
            }}
            
            /* Style pour la barre d'outils */
            QToolBar {{
                background-color: #{COLORS["PRIMARY"]};
                border: none;
                spacing: 10px;
                padding: 5px;
            }}
            
            QToolButton {{
                background-color: transparent;
                color: white;
                border-radius: 4px;
                padding: 5px;
            }}
            
            QToolButton:hover {{
                background-color: #{COLORS["PRIMARY_DARK"]};
            }}
            
            /* Style pour les onglets */
            QTabWidget::pane {{
                border: 1px solid #{COLORS["PRIMARY"]};
                border-radius: 4px;
                background-color: white;
            }}
            
            QTabBar::tab {{
                background-color: #{COLORS["PRIMARY"]};
                color: white;
                border: 1px solid #{COLORS["PRIMARY_DARK"]};
                padding: 8px 15px;
                margin-right: 2px;
            }}
            
            QTabBar::tab:selected {{
                background-color: #{COLORS["PRIMARY_DARK"]};
                color: white;
            }}
            
            QTabBar::tab:hover {{
                background-color: #{COLORS["PRIMARY_LIGHT"]};
                color: white;
            }}
            
            /* Style pour les groupes et bordures */
            QGroupBox {{
                font-weight: bold;
                border: 2px solid #{COLORS["PRIMARY"]};
                border-radius: 8px;
                margin-top: 12px;
                padding: 15px;
                background-color: white;
            }}
            
            QGroupBox::title {{
                color: #{COLORS["PRIMARY"]};
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }}
            
            /* Style pour les boutons */
            QPushButton {{
                background-color: #{COLORS["PRIMARY"]};
                color: white;
                padding: 8px 15px;
                border-radius: 4px;
                font-weight: bold;
                min-height: 30px;
            }}
            
            QPushButton:hover {{
                background-color: #{COLORS["PRIMARY_DARK"]};
            }}
            
            QPushButton:disabled {{
                background-color: #a0a0a0;
                color: #d0d0d0;
            }}
            
            /* Style pour les tableaux */
            QTableWidget {{
                border: 1px solid #{COLORS["PRIMARY"]};
                gridline-color: #{COLORS["PRIMARY"]};
            }}
            
            QTableWidget QHeaderView::section {{
                background-color: #{COLORS["PRIMARY"]};
                color: white;
                padding: 5px;
                border: 1px solid white;
            }}
            
            /* Style pour la barre de progression */
            QProgressBar {{
                border: 1px solid #{COLORS["PRIMARY"]};
                border-radius: 4px;
                text-align: center;
            }}
            
            QProgressBar::chunk {{
                background-color: #{COLORS["PRIMARY"]};
                width: 10px;
                margin: 0.5px;
            }}
            
            /* Style pour les checkboxes */
            QCheckBox::indicator:checked {{
                background-color: #{COLORS["PRIMARY"]};
                border: 1px solid #{COLORS["PRIMARY_DARK"]};
            }}
            
            /* Style pour les étiquettes de titre */
            QLabel[accessibleName="title"] {{
                color: #{COLORS["PRIMARY"]};
                font-weight: bold;
            }}
            
            /* Style pour les en-têtes de vue */
            QHeaderView::section {{
                background-color: #{COLORS["PRIMARY"]};
                color: white;
            }}
            
            /* Style pour tous les bordures et contours */
            * {{
                border-color: #{COLORS["PRIMARY"]};
            }}
        """)
    
    def create_menu(self):
        """Crée le menu de l'application."""
        menubar = self.menuBar()
        
        # Menu Fichier
        file_menu = menubar.addMenu(self.translate("file_menu"))
        
        # Action Ouvrir un répertoire
        open_action = QAction(self.translate("open_dir"), self)
        open_action.setShortcut(QKeySequence.StandardKey.Open)
        open_action.triggered.connect(self.choose_directory)
        file_menu.addAction(open_action)
        
        file_menu.addSeparator()
        
        # Action Enregistrer les paramètres
        save_settings_action = QAction(self.translate("save_settings"), self)
        save_settings_action.triggered.connect(self.save_settings)
        file_menu.addAction(save_settings_action)
        
        # Action Charger les paramètres
        load_settings_action = QAction(self.translate("load_settings"), self)
        load_settings_action.triggered.connect(self.load_settings)
        file_menu.addAction(load_settings_action)
        
        file_menu.addSeparator()
        
        # Action Quitter
        exit_action = QAction(self.translate("exit"), self)
        exit_action.setShortcut(QKeySequence.StandardKey.Quit)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Menu Édition
        edit_menu = menubar.addMenu(self.translate("edit_menu"))
        
        # Action Sélectionner tout
        select_all_action = QAction(self.translate("select_all"), self)
        select_all_action.triggered.connect(self.select_all_files)
        edit_menu.addAction(select_all_action)
        
        # Action Désélectionner tout
        deselect_all_action = QAction(self.translate("deselect_all"), self)
        deselect_all_action.triggered.connect(self.deselect_all_files)
        edit_menu.addAction(deselect_all_action)
        
        # Action Inverser la sélection
        invert_selection_action = QAction(self.translate("invert_selection"), self)
        invert_selection_action.triggered.connect(self.invert_file_selection)
        edit_menu.addAction(invert_selection_action)
        
        # Menu Outils
        tools_menu = menubar.addMenu(self.translate("tools_menu"))
        
        # Action Aperçu des données
        preview_action = QAction(self.translate("preview_tool"), self)
        preview_action.triggered.connect(self.show_preview)
        tools_menu.addAction(preview_action)
        
        # Action Format de date
        date_format_action = QAction(self.translate("date_format"), self)
        date_format_action.triggered.connect(self.show_date_format_dialog)
        tools_menu.addAction(date_format_action)
        
        # Menu Langue
        language_menu = menubar.addMenu(self.translate("language_menu"))
        
        # Action Changer la langue
        change_language_action = QAction(self.translate("language"), self)
        change_language_action.triggered.connect(self.show_language_dialog)
        language_menu.addAction(change_language_action)
        
        # Menu Aide
        about_menu = menubar.addMenu(self.translate("about"))
        
        # Action Aide
        about_action = QAction(self.translate("about"), self)
        about_action.triggered.connect(self.show_about_dialog)
        about_menu.addAction(about_action)
    
    def create_toolbar(self):
        """Crée la barre d'outils de l'application."""
        toolbar = QToolBar(self)
        toolbar.setMovable(False)
        toolbar.setIconSize(QSize(32, 32))
        
        # Action Ouvrir un répertoire
        open_action = QAction(self.style().standardIcon(QStyle.StandardPixmap.SP_DirOpenIcon), self.translate("open_dir"), self)
        open_action.triggered.connect(self.choose_directory)
        toolbar.addAction(open_action)
        
        toolbar.addSeparator()
        
        # Action Aperçu des données
        preview_action = QAction(self.style().standardIcon(QStyle.StandardPixmap.SP_FileDialogContentsView), self.translate("preview_tool"), self)
        preview_action.triggered.connect(self.show_preview)
        toolbar.addAction(preview_action)
        
        # Action Compiler
        compile_action = QAction(self.style().standardIcon(QStyle.StandardPixmap.SP_MediaPlay), self.translate("compile_tool"), self)
        compile_action.triggered.connect(self.compile_files)
        toolbar.addAction(compile_action)
        
        self.addToolBar(toolbar)
    
    def create_main_layout(self):
        """Crée la mise en page principale de l'application."""
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)
        
        # Création des onglets
        self.tabs = QTabWidget()
        self.create_compilation_tab()
        self.create_advanced_options_tab()
        self.create_date_format_tab()
        self.create_preview_tab()
        self.create_help_tab()
        self.create_about_tab()
        
        # Appliquer un style spécifique à la barre d'onglets
        self.tabs.setStyleSheet(f"""
            QTabWidget::pane {{
                border: 1px solid #{COLORS["PRIMARY"]};
            }}
            QTabBar {{
                background-color: #{COLORS["PRIMARY"]};
            }}
            QTabBar::tab {{
                background-color: #{COLORS["PRIMARY"]};
                color: white;
                padding: 8px 15px;
            }}
            QTabBar::tab:selected {{
                background-color: #{COLORS["PRIMARY_DARK"]};
                font-weight: bold;
            }}
            QTabBar::tab:!selected {{
                margin-top: 2px;
            }}
        """)
        
        main_layout.addWidget(self.tabs)
    

    def create_compilation_tab(self):
        """Crée l'onglet principal de compilation."""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Groupe sélection des fichiers
        files_group = self.create_files_group()
        layout.addWidget(files_group)
        
        # Groupe options de compilation
        options_group = self.create_options_group()
        layout.addWidget(options_group)
        
        # Barre de progression
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Bouton de compilation
        compile_layout = QHBoxLayout()
        self.button_compile = QPushButton(self.translate("start_compilation"))
        self.button_compile.setFont(QFont("Segoe UI", FONT_SIZES["LARGE"]))
        self.button_compile.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MediaPlay))
        self.button_compile.setStyleSheet(f"background-color: #{COLORS['PRIMARY']}; color: white;")
        compile_layout.addStretch()
        compile_layout.addWidget(self.button_compile)
        compile_layout.addStretch()
        layout.addLayout(compile_layout)
        
        # Status label
        self.status_label = QLabel("")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.status_label)
        
        tab.setLayout(layout)
        self.tabs.addTab(tab, self.translate("compilation_options"))
    
    def create_files_group(self):
        """Crée le groupe d'éléments pour la sélection des fichiers."""
        group = QGroupBox(self.translate("file_selection"))
        layout = QVBoxLayout()
        
        # Date et heure
        self.datetime_label = QLabel()
        self.update_datetime()
        
        # Timer pour mettre à jour la date et l'heure
        self.datetime_timer = QTimer()
        self.datetime_timer.timeout.connect(self.update_datetime)
        self.datetime_timer.start(1000)  # Mise à jour toutes les secondes
        
        # Sélection du répertoire
        dir_layout = QHBoxLayout()
        self.label_directory = QLabel(self.translate("no_directory"))
        self.button_choose_directory = QPushButton(self.translate("choose_directory"))
        self.button_choose_directory.setFont(QFont("Segoe UI", FONT_SIZES["NORMAL"]))
        self.button_choose_directory.setIcon(QIcon(resource_path("folder.ico")))
        self.button_choose_directory.setStyleSheet(f"background-color: #{COLORS['PRIMARY']}; color: white;")

        dir_layout.addWidget(self.label_directory)
        dir_layout.addWidget(self.button_choose_directory)
        layout.addLayout(dir_layout)
        
        # Liste des fichiers
        self.list_files = QListWidget()
        self.list_files.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        
        # Options de sélection
        selection_layout = QHBoxLayout()
        self.checkbox_all_files = QCheckBox(self.translate("select_all_files"))
        self.label_file_count = QLabel(self.translate("files_selected", 0))
        selection_layout.addWidget(self.checkbox_all_files)
        selection_layout.addStretch()
        selection_layout.addWidget(self.label_file_count)
        selection_layout.addStretch()
        selection_layout.addWidget(self.datetime_label)
        
        layout.addLayout(selection_layout)
        layout.addWidget(self.list_files)
        group.setLayout(layout)
        return group
    
    def create_options_group(self):
        """Crée le groupe d'éléments pour les options de compilation."""
        group = QGroupBox(self.translate("compilation_options"))
        grid_layout = QGridLayout()
        
        # Première colonne : Spinboxes et labels
        # Ligne 0
        self.label_header_start = QLabel(self.translate("header_start_row"))
        self.spinbox_header_start = QSpinBox()
        self.spinbox_header_start.setMinimum(1)
        self.spinbox_header_start.setMaximum(20)
        self.spinbox_header_start.setValue(1)
        header_start_container = QHBoxLayout()
        header_start_container.addWidget(self.label_header_start)
        header_start_container.addWidget(self.spinbox_header_start)
        header_start_container.addStretch()
        
        # Ligne 1
        self.label_header = QLabel(self.translate("header_rows"))
        self.spinbox_header = QSpinBox()
        self.spinbox_header.setMinimum(1)
        self.spinbox_header.setMaximum(15)
        self.spinbox_header.setValue(1)
        header_container = QHBoxLayout()
        header_container.addWidget(self.label_header)
        header_container.addWidget(self.spinbox_header)
        header_container.addStretch()
        
        # Deuxième colonne : Checkboxes pour répéter et fusionner
        self.checkbox_repeat_header = QCheckBox(self.translate("repeat_headers"))
        self.checkbox_merge_headers = QCheckBox(self.translate("merge_headers"))
        
        # Troisième colonne : Nom de fichier et sortie
        self.checkbox_add_filename = QCheckBox(self.translate("add_filename"))
        self.checkbox_add_filename.setChecked(True)
        
        self.label_output_name = QLabel(self.translate("output_filename"))
        self.lineedit_output_name = QLineEdit("compilation.xlsx")
        output_container = QHBoxLayout()
        output_container.addWidget(self.label_output_name)
        output_container.addWidget(self.lineedit_output_name)
        
        # Option de vérification préliminaire
        verification_layout = QHBoxLayout()
        self.checkbox_verify_files = QCheckBox(self.translate("enable_verification"))
        self.checkbox_verify_files.setChecked(True)
        verification_layout.addWidget(self.checkbox_verify_files)
        verification_layout.addStretch()
        
        # Ajout des widgets au QGridLayout
        # Première colonne (0)
        grid_layout.addLayout(header_start_container, 0, 0)
        grid_layout.addLayout(header_container, 1, 0)
        
        # Deuxième colonne (1)
        grid_layout.addWidget(self.checkbox_repeat_header, 0, 1)
        grid_layout.addWidget(self.checkbox_merge_headers, 1, 1)
        
        # Troisième colonne (2)
        grid_layout.addWidget(self.checkbox_add_filename, 0, 2)
        grid_layout.addLayout(output_container, 1, 2)
        
        # Ligne supplémentaire pour la vérification
        grid_layout.addLayout(verification_layout, 2, 0, 1, 3)
        
        # Définir l'espacement et les marges
        grid_layout.setSpacing(20)  # Espacement entre les éléments
        grid_layout.setContentsMargins(20, 20, 20, 20)  # Marges autour de la grille
        
        # Définir les colonnes pour qu'elles aient la même largeur
        grid_layout.setColumnStretch(0, 1)
        grid_layout.setColumnStretch(1, 1)
        grid_layout.setColumnStretch(2, 1)
        
        # Alignement vertical des éléments
        grid_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        
        group.setLayout(grid_layout)
        return group
    
    def create_advanced_options_tab(self):
        """Crée l'onglet des options avancées."""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Groupe traitement des données
        data_group = QGroupBox(self.translate("compilation_options"))
        data_layout = QVBoxLayout()
        
        self.checkbox_remove_duplicates = QCheckBox(self.translate("remove_duplicates"))
        self.checkbox_remove_duplicates.setChecked(True)
        
        self.checkbox_remove_empty_rows = QCheckBox(self.translate("remove_empty_rows"))
        self.checkbox_remove_empty_rows.setChecked(False)
        
        sort_layout = QVBoxLayout()
        self.checkbox_sort_data = QCheckBox(self.translate("sort_data"))
        
        sort_options = QHBoxLayout()
        self.label_sort_column = QLabel(self.translate("sort_column"))
        self.lineedit_sort_column = QLineEdit("A")
        self.lineedit_sort_column.setEnabled(False)
        sort_options.addWidget(self.label_sort_column)
        sort_options.addWidget(self.lineedit_sort_column)
        sort_options.addStretch()
        
        sort_layout.addWidget(self.checkbox_sort_data)
        sort_layout.addLayout(sort_options)
        
        data_layout.addWidget(self.checkbox_remove_duplicates)
        data_layout.addWidget(self.checkbox_remove_empty_rows)
        data_layout.addLayout(sort_layout)
        data_group.setLayout(data_layout)
        
        # Groupe formatage
        format_group = QGroupBox(self.translate("compilation_options"))
        format_layout = QVBoxLayout()
        
        self.checkbox_auto_width = QCheckBox(self.translate("auto_width"))
        self.checkbox_auto_width.setChecked(True)
        
        self.checkbox_freeze_header = QCheckBox(self.translate("freeze_headers"))
        self.checkbox_freeze_header.setChecked(True)
        
        format_layout.addWidget(self.checkbox_auto_width)
        format_layout.addWidget(self.checkbox_freeze_header)
        format_group.setLayout(format_layout)
        
        # Groupe formats des fichiers
        format_files_group = QGroupBox(self.translate("file_formats"))
        format_files_layout = QVBoxLayout()
        
        self.checkbox_excel = QCheckBox(self.translate("excel_files"))
        self.checkbox_excel.setChecked(True)
        self.checkbox_excel.setEnabled(False)  # Toujours activé
        
        self.checkbox_csv = QCheckBox(self.translate("csv_files"))
        self.checkbox_csv.setChecked(True)
        
        format_files_layout.addWidget(self.checkbox_excel)
        format_files_layout.addWidget(self.checkbox_csv)
        format_files_group.setLayout(format_files_layout)
        
        layout.addWidget(data_group)
        layout.addWidget(format_group)
        layout.addWidget(format_files_group)
        layout.addStretch()
        tab.setLayout(layout)
        self.tabs.addTab(tab, self.translate("advanced_options"))
    
    def create_date_format_tab(self):
        """Crée l'onglet de format de date."""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Groupe format de date
        group = QGroupBox(self.translate("date_format_options"))
        group_layout = QVBoxLayout()
        
        # Créer les boutons radio pour chaque format prédéfini
        self.date_radio_group = QButtonGroup(self)
        formats = [
            ("STANDARD", "date_format_standard"),
            ("FRENCH", "date_format_french"),
            ("US", "date_format_us"),
            ("DATETIME", "date_format_datetime"),
            ("DATETIME_FRENCH", "date_format_datetime_french"),
            ("DATE_ONLY", "date_format_date_only"),
            ("TIME_ONLY", "date_format_time_only"),
            ("SHORT", "date_format_short"),
            ("CUSTOM", "date_format_custom")
        ]
        
        self.date_radio_buttons = {}
        
        for i, (format_key, label_key) in enumerate(formats):
            radio = QRadioButton(self.translate(label_key))
            self.date_radio_group.addButton(radio, i)
            self.date_radio_buttons[format_key] = radio
            
            if format_key == "CUSTOM":
                custom_layout = QHBoxLayout()
                custom_layout.addWidget(radio)
                self.date_custom_edit = QLineEdit()
                self.date_custom_edit.setPlaceholderText("dd/MM/yyyy HH:mm:ss")
                self.date_custom_edit.setEnabled(False)
                custom_layout.addWidget(self.date_custom_edit)
                group_layout.addLayout(custom_layout)
            else:
                group_layout.addWidget(radio)
        
        group.setLayout(group_layout)
        layout.addWidget(group)
        
        # Exemple avec la date actuelle
        example_layout = QHBoxLayout()
        example_layout.addWidget(QLabel(self.translate("preview") + ":"))
        self.date_example_label = QLabel()
        self.update_date_example()
        example_layout.addWidget(self.date_example_label)
        layout.addLayout(example_layout)
        
        # Bouton Appliquer
        button_layout = QHBoxLayout()
        apply_button = QPushButton(self.translate("apply"))
        apply_button.clicked.connect(self.apply_date_format)
        button_layout.addStretch()
        button_layout.addWidget(apply_button)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        layout.addStretch()
        
        tab.setLayout(layout)
        self.tabs.addTab(tab, self.translate("date_format"))
        
        # Connecter les signaux
        self.date_radio_group.buttonClicked.connect(self.on_date_format_changed)
        self.date_custom_edit.textChanged.connect(self.on_date_custom_format_changed)
        
        # Sélectionner le format actuel
        if self.date_format in self.date_radio_buttons:
            self.date_radio_buttons[self.date_format].setChecked(True)
            if self.date_format == "CUSTOM":
                self.date_custom_edit.setEnabled(True)
                self.date_custom_edit.setText(DATE_FORMATS["CUSTOM"]["format"])
    
    def create_preview_tab(self):
        """Crée l'onglet de prévisualisation des données."""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Options de prévisualisation
        options_layout = QHBoxLayout()
        
        # Sélection du fichier
        file_label = QLabel(self.translate("filename") + ":")
        self.preview_combo = QComboBox()
        self.preview_combo.currentIndexChanged.connect(self.on_preview_file_changed)
        
        # Bouton Actualiser
        self.preview_refresh_button = QPushButton(self.translate("refresh_preview"))
        self.preview_refresh_button.clicked.connect(self.refresh_preview)
        
        options_layout.addWidget(file_label)
        options_layout.addWidget(self.preview_combo, 1)
        options_layout.addWidget(self.preview_refresh_button)
        
        layout.addLayout(options_layout)
        
        # Tableau de prévisualisation
        self.preview_table = QTableWidget()
        self.preview_table.setAlternatingRowColors(True)
        layout.addWidget(self.preview_table)
        
        # Informations sur la prévisualisation
        self.preview_info_label = QLabel(self.translate("preview_limited", 200))
        layout.addWidget(self.preview_info_label)
        
        tab.setLayout(layout)
        self.tabs.addTab(tab, self.translate("preview"))

    def create_help_tab(self):
        """Crée l'onglet d'aide."""
        tab = QWidget()
        layout = QVBoxLayout()
        
        help_text = """
        <style>
        body {
            font-family: "Segoe UI", sans-serif;
            margin: 0 20px;
            line-height: 1.6;
            color: #333;
        }

        h2, h3 {
            color: #2e7d32;
            font-weight: bold;
            margin: 25px 0 15px 0;
        }

        h4 {
            margin: 12px 0;
            padding-left: 20px;
            font-weight: normal;
        }

        ul {
            margin-left: 50px;
            line-height: 1.5;
        }

        li {
            margin: 15px 0;
        }

        .section-content {
            padding-left: 30px;
        }

        b {
            color: #1b5e20;
        }

        .highlight {
            background-color: #e8f5e9;
            padding: 2px 5px;
            border-radius: 3px;
        }
        </style>

        <h2>Guide d'utilisation du Compilateur Excel</h2>
        
        <h3>I. Sélection des fichiers</h3>
        <div class="section-content">
            <h4>• Cliquez sur <b>"Choisir un répertoire"</b> pour sélectionner le dossier contenant vos fichiers Excel et CSV</h4>
            <h4>• Sélectionnez les fichiers à compiler dans la liste</h4>
            <h4>• Utilisez la case <b>"Sélectionner tous les fichiers"</b> pour tout sélectionner/désélectionner</h4>
            <h4>• Le nombre de fichiers sélectionnés est affiché en temps réel</h4>
        </div>
        
        <h3>II. Options de compilation</h3>
        <div class="section-content">
            <h4>• <b>Ligne de début de l'en-tête</b> : Spécifiez à quelle ligne commence l'en-tête dans vos fichiers</h4> 
            <h4>• <b>Nombre de lignes d'en-tête</b> : Indiquez combien de lignes constituent l'en-tête</h4> 
            <h4>• <b>Répéter les en-têtes</b> : Réinsère les en-têtes entre chaque fichier dans la compilation</h4> 
            <h4>• <b>Fusionner les en-têtes multi-niveaux</b> : Conserve la fusion des cellules d'en-tête</h4> 
            <h4>• <b>Ajouter les noms des fichiers sources</b> : Ajoute une colonne avec les noms des fichiers d'origine</h4> 
            <h4>• <b>Nom du fichier de sortie</b> : Définissez le nom du fichier compilé (.xlsx sera ajouté automatiquement)</h4>
            <h4>• <b>Vérification préliminaire</b> : Active ou désactive l'analyse des fichiers avant compilation</h4>
        </div>
        
        <h3>III. Options avancées</h3>
        <div class="section-content">
            <h4><u>Traitement des données :</u></h4>
            <h4>• <b>Supprimer les doublons</b> : Élimine les lignes identiques</h4> 
            <h4>• <b>Supprimer les lignes vides</b> : Retire les lignes ne contenant aucune donnée</h4> 
            <h4>• <b>Trier les données</b> : Trie le contenu selon une colonne spécifique</h4> 
            <h4>• <b>Colonne de tri</b> : Spécifiez la colonne pour le tri (ex: A pour première colonne)</h4> 
            
            <h4><u>Formatage :</u></h4>
            <h4>• <b>Ajuster la largeur des colonnes</b> : Adapte automatiquement la largeur selon le contenu</h4> 
            <h4>• <b>Figer les en-têtes</b> : Maintient l'en-tête visible lors du défilement</h4>
            
            <h4><u>Formats de fichiers supportés :</u></h4>
            <h4>• <b>Excel</b> : Traite les fichiers .xlsx et .xls</h4>
            <h4>• <b>CSV</b> : Traite les fichiers .csv avec détection automatique du délimiteur</h4>
        </div>
        
        <h3>IV. Format de date</h3>
        <div class="section-content">
            <h4>• Sélectionnez le format de date à utiliser pour les cellules contenant des dates</h4>
            <h4>• Formats prédéfinis disponibles: standard, français, américain, etc.</h4>
            <h4>• Option de format personnalisé pour des besoins spécifiques</h4>
            <h4>• L'aperçu montre comment la date actuelle apparaîtra dans ce format</h4>
        </div>
        
        <h3>V. Prévisualisation</h3>
        <div class="section-content">
            <h4>• Examine le contenu des fichiers avant de les compiler</h4>
            <h4>• Sélectionnez un fichier dans la liste déroulante pour le prévisualiser</h4>
            <h4>• Vérifie la structure des données, les en-têtes, etc.</h4>
            <h4>• L'aperçu est limité aux 200 premières lignes pour des raisons de performance</h4>
        </div>
        
        <h3>VI. Internationalisation</h3>
        <div class="section-content">
            <h4>• L'application est disponible en français et en anglais</h4>
            <h4>• Changez la langue via le menu Langue</h4>
            <h4>• La langue choisie est sauvegardée dans les préférences</h4>
        </div>
        
        <h3>VII. Vérification préliminaire des fichiers</h3>
        <div class="section-content">
            <h4>• Avant la compilation, l'application analyse tous les fichiers sélectionnés</h4>
            <h4>• Un rapport détaillé identifie les fichiers compatibles et incompatibles</h4>
            <h4>• Pour chaque fichier problématique, un motif précis est affiché</h4>
            <h4>• Vous pouvez choisir d'ignorer les fichiers incompatibles ou d'annuler la compilation</h4>
        </div>
        
        <h3>VIII. Bonnes pratiques</h3>
        <div class="section-content">
            <h4>• Faites une sauvegarde de vos fichiers avant la compilation</h4>
            <h4>• Vérifiez que tous les fichiers ont une structure similaire</h4>
            <h4>• Pour les gros fichiers, traitez-les par lots</h4>
            <h4>• Utilisez des noms explicites pour les fichiers de sortie</h4>
            <h4>• Activez la vérification préliminaire pour éviter les erreurs</h4>
            <h4>• Prévisualisez les données avant compilation pour vérifier leur structure</h4>
        </div>
        """

        # Création de l'étiquette avec le texte d'aide
        help_label = QLabel(help_text)
        help_label.setTextFormat(Qt.TextFormat.RichText)
        help_label.setWordWrap(True)
        
        # Ajout d'un ScrollArea pour permettre le défilement
        scroll = QScrollArea()
        scroll.setWidget(help_label)
        scroll.setWidgetResizable(True)
        
        layout.addWidget(scroll)
        tab.setLayout(layout)
        self.tabs.addTab(tab, self.translate("help"))
    
    def create_about_tab(self):
        """Crée l'onglet À propos."""
        tab = QWidget()
        layout = QVBoxLayout()
        
        about_text = f"""
        <div style="text-align: center; margin: 50px 20px;">
            <h1 style="color: #{COLORS['PRIMARY']}; margin-bottom: 30px;">{self.translate("app_title")}</h1>
            <h2>Version 3.0</h2>
            <p style="font-size: 16px; margin: 30px 0;">
                Développé par:<br>
                <strong style="font-size: 20px; color: #{COLORS['PRIMARY_DARK']};">{self.translate("developer_name")}</strong><br>
                {self.translate("developer_title")}<br>
            <h2>Email : zimkada@gmail.com</h2>
            
            </p>
            <p style="margin-top: 40px; color: #555; font-style: italic;">
                Copyright © 2025. {self.translate("copyright_notice")}
            </p>
        </div>
        """
        
        label = QLabel(about_text)
        label.setTextFormat(Qt.TextFormat.RichText)
        layout.addWidget(label)
        
        tab.setLayout(layout)
        self.tabs.addTab(tab, self.translate("about"))
    
    def connect_signals(self):
        """Connecte les signaux aux slots."""
        # Boutons
        self.button_choose_directory.clicked.connect(self.choose_directory)
        self.button_compile.clicked.connect(self.compile_files)
        
        # Changements de sélection
        self.list_files.itemSelectionChanged.connect(self.update_selection_count)
        self.checkbox_all_files.stateChanged.connect(self.toggle_all_files)
        
        # Options avancées
        self.checkbox_sort_data.stateChanged.connect(self.toggle_sort_options)
        
        # Option de vérification préliminaire
        self.checkbox_verify_files.stateChanged.connect(self.toggle_verification)
        
        # Format de date
        self.date_radio_group.buttonClicked.connect(self.on_date_format_changed)
        self.date_custom_edit.textChanged.connect(self.on_date_custom_format_changed)
    
    def update_ui_language(self):
        """Met à jour la langue de l'interface utilisateur."""
        # Mettre à jour le titre de la fenêtre
        self.setWindowTitle(self.translate("app_title"))
        
        # Mettre à jour les onglets
        self.tabs.setTabText(0, self.translate("compilation_options"))
        self.tabs.setTabText(1, self.translate("advanced_options"))
        self.tabs.setTabText(2, self.translate("date_format"))
        self.tabs.setTabText(3, self.translate("preview"))
        self.tabs.setTabText(4, self.translate("help"))
        self.tabs.setTabText(5, self.translate("about"))
        
        # Mettre à jour les groupes
        for group in self.findChildren(QGroupBox):
            if group.title() == "Sélection des fichiers" or group.title() == "File Selection":
                group.setTitle(self.translate("file_selection"))
            elif group.title() == "Options de compilation" or group.title() == "Compilation Options":
                group.setTitle(self.translate("compilation_options"))
            elif group.title() == "Formats de fichiers supportés" or group.title() == "Supported File Formats":
                group.setTitle(self.translate("file_formats"))
            elif group.title().startswith("Options de format de date") or group.title().startswith("Date Format Options"):
                group.setTitle(self.translate("date_format_options"))
        
        # Mettre à jour les labels
        self.label_directory.setText(self.translate("no_directory") if not self.directory else self.directory)
        self.label_header_start.setText(self.translate("header_start_row"))
        self.label_header.setText(self.translate("header_rows"))
        self.label_output_name.setText(self.translate("output_filename"))
        self.label_sort_column.setText(self.translate("sort_column"))
        self.update_datetime()
        self.update_selection_count()
        
        # Mettre à jour les checkboxes
        self.checkbox_all_files.setText(self.translate("select_all_files"))
        self.checkbox_repeat_header.setText(self.translate("repeat_headers"))
        self.checkbox_merge_headers.setText(self.translate("merge_headers"))
        self.checkbox_add_filename.setText(self.translate("add_filename"))
        self.checkbox_verify_files.setText(self.translate("enable_verification"))
        self.checkbox_remove_duplicates.setText(self.translate("remove_duplicates"))
        self.checkbox_remove_empty_rows.setText(self.translate("remove_empty_rows"))
        self.checkbox_sort_data.setText(self.translate("sort_data"))
        self.checkbox_auto_width.setText(self.translate("auto_width"))
        self.checkbox_freeze_header.setText(self.translate("freeze_headers"))
        self.checkbox_excel.setText(self.translate("excel_files"))
        self.checkbox_csv.setText(self.translate("csv_files"))
        
        # Mettre à jour les boutons
        self.button_choose_directory.setText(self.translate("choose_directory"))
        self.button_compile.setText(self.translate("start_compilation"))
        self.preview_refresh_button.setText(self.translate("refresh_preview"))
        
        # Mettre à jour les boutons radio du format de date
        for format_key, label_key in [
            ("STANDARD", "date_format_standard"),
            ("FRENCH", "date_format_french"),
            ("US", "date_format_us"),
            ("DATETIME", "date_format_datetime"),
            ("DATETIME_FRENCH", "date_format_datetime_french"),
            ("DATE_ONLY", "date_format_date_only"),
            ("TIME_ONLY", "date_format_time_only"),
            ("SHORT", "date_format_short"),
            ("CUSTOM", "date_format_custom")
        ]:
            if format_key in self.date_radio_buttons:
                self.date_radio_buttons[format_key].setText(self.translate(label_key))
        
        # Recréer le menu
        menubar = self.menuBar()
        menubar.clear()
        self.create_menu()
        
        # Mettre à jour les infos de prévisualisation
        self.preview_info_label.setText(self.translate("preview_limited", 200))
        
        # Recréer la barre d'outils
        for toolbar in self.findChildren(QToolBar):
            self.removeToolBar(toolbar)
        self.create_toolbar()
        
        # Recharger les onglets d'aide et à propos
        self.tabs.removeTab(5)  # À propos
        self.tabs.removeTab(4)  # Aide
        self.create_help_tab()
        self.create_about_tab()
    
    def update_datetime(self):
        """Met à jour l'affichage de la date et de l'heure."""
        current_datetime = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        self.datetime_label.setText(self.translate("date_time", current_datetime))
    
    def toggle_verification(self, state):
        """Active ou désactive la vérification préliminaire des fichiers."""
        self.verification_enabled = state == Qt.CheckState.Checked.value
    
    def toggle_all_files(self, state):
        """Sélectionne ou désélectionne tous les fichiers."""
        for i in range(self.list_files.count()):
            self.list_files.item(i).setSelected(state == Qt.CheckState.Checked.value)
    
    def select_all_files(self):
        """Sélectionne tous les fichiers."""
        self.checkbox_all_files.setChecked(True)
    
    def deselect_all_files(self):
        """Désélectionne tous les fichiers."""
        self.checkbox_all_files.setChecked(False)
    
    def invert_file_selection(self):
        """Inverse la sélection des fichiers."""
        for i in range(self.list_files.count()):
            item = self.list_files.item(i)
            item.setSelected(not item.isSelected())
    
    def update_selection_count(self):
        """Met à jour le compteur de fichiers sélectionnés."""
        selected_count = len(self.list_files.selectedItems())
        self.label_file_count.setText(self.translate("files_selected", selected_count))
        self.button_compile.setEnabled(selected_count > 0)
    
    def toggle_sort_options(self, state):
        """Active ou désactive les options de tri."""
        self.lineedit_sort_column.setEnabled(state == Qt.CheckState.Checked.value)
    
    def on_date_format_changed(self, button):
        """
        Appelé lorsque l'utilisateur change de format de date.
        
        Args:
            button: Bouton radio sélectionné
        """
        for format_key, radio in self.date_radio_buttons.items():
            if radio == button:
                self.date_custom_edit.setEnabled(format_key == "CUSTOM")
                break
        
        self.update_date_example()
    
    def on_date_custom_format_changed(self, text):
        """
        Appelé lorsque l'utilisateur modifie le format personnalisé.
        
        Args:
            text: Nouveau texte du format personnalisé
        """
        DATE_FORMATS["CUSTOM"]["format"] = text
        self.update_date_example()
    
    def update_date_example(self):
        """Met à jour l'exemple de format de date."""
        now = datetime.now()
        
        for format_key, radio in self.date_radio_buttons.items():
            if radio.isChecked():
                if format_key == "CUSTOM":
                    try:
                        # Convertir le format personnalisé en format de date Python
                        user_format = self.date_custom_edit.text()
                        # Remplacer les tokens de format
                        py_format = user_format.replace("dd", "%d").replace("MM", "%m").replace("yyyy", "%Y")
                        py_format = py_format.replace("HH", "%H").replace("mm", "%M").replace("ss", "%S")
                        py_format = py_format.replace("yy", "%y")
                        
                        formatted_date = now.strftime(py_format)
                        self.date_example_label.setText(formatted_date)
                    except Exception:
                        self.date_example_label.setText("Format invalide")
                else:
                    # Utiliser le format prédéfini
                    date_format = DATE_FORMATS[format_key]["format"]
                    
                    # Convertir en format Python
                    py_format = date_format.replace("dd", "%d").replace("MM", "%m").replace("yyyy", "%Y")
                    py_format = py_format.replace("HH", "%H").replace("mm", "%M").replace("ss", "%S")
                    py_format = py_format.replace("yy", "%y")
                    
                    formatted_date = now.strftime(py_format)
                    self.date_example_label.setText(formatted_date)
                break
    
    def apply_date_format(self):
        """Applique le format de date sélectionné."""
        for format_key, radio in self.date_radio_buttons.items():
            if radio.isChecked():
                self.date_format = format_key
                
                if format_key == "CUSTOM":
                    # Enregistrer le format personnalisé
                    custom_format = self.date_custom_edit.text()
                    DATE_FORMATS["CUSTOM"]["format"] = custom_format
                    
                    # Générer un format Excel personnalisé
                    excel_format = custom_format
                    excel_format = excel_format.replace("dd", "dd").replace("MM", "mm").replace("yyyy", "yyyy")
                    excel_format = excel_format.replace("HH", "hh").replace("mm", "mm").replace("ss", "ss")
                    excel_format = excel_format.replace("yy", "yy")
                    
                    DATE_FORMATS["CUSTOM"]["excel_format"] = excel_format
                
                QMessageBox.information(
                    self,
                    self.translate("date_format"),
                    self.translate("settings_saved")
                )
                break
    
    def on_preview_file_changed(self, index):
        """
        Appelé lorsque l'utilisateur change de fichier dans le combobox de prévisualisation.
        
        Args:
            index: Index du fichier sélectionné
        """
        if index >= 0:
            self.refresh_preview()

    def refresh_preview(self):
        """Actualise la prévisualisation du fichier sélectionné."""
        if not self.directory or self.preview_combo.count() == 0:
            return
        
        current_index = self.preview_combo.currentIndex()
        if current_index < 0:
            return
        
        selected_file = self.preview_combo.itemText(current_index)
        file_path = os.path.join(self.directory, selected_file)
        
        try:
            if selected_file.lower().endswith(('.xlsx', '.xls')):
                self.load_excel_preview_tab(file_path)
            elif selected_file.lower().endswith('.csv'):
                self.load_csv_preview_tab(file_path)
        except Exception as e:
            QMessageBox.warning(
                self,
                self.translate("error"),
                f"{self.translate('error')}: {str(e)}"
            )
    
    def load_excel_preview_tab(self, file_path):
        """
        Charge la prévisualisation d'un fichier Excel dans l'onglet prévisualisation.
        
        Args:
            file_path: Chemin du fichier Excel à prévisualiser
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active
            
            # Obtenir les données (limité à 200 lignes)
            preview_data = []
            headers = []
            
            # Récupérer les en-têtes
            header_start = self.spinbox_header_start.value()
            header_rows = self.spinbox_header.value()
            
            for row in range(header_start, header_start + header_rows):
                header_row = []
                for cell in ws[row]:
                    header_row.append(cell.value)
                headers.append(header_row)
            
            # Récupérer les données
            row_count = 0
            for row in ws.iter_rows(min_row=header_start + header_rows):
                if row_count >= 200:
                    break
                
                row_data = [cell.value for cell in row]
                preview_data.append(row_data)
                row_count += 1
            
            # Afficher les données
            self.display_preview_tab(headers, preview_data)
            
            wb.close()
        except Exception as e:
            raise ValueError(f"Erreur lors de la lecture du fichier Excel: {str(e)}")
    
    def load_csv_preview_tab(self, file_path):
        """
        Charge la prévisualisation d'un fichier CSV dans l'onglet prévisualisation.
        
        Args:
            file_path: Chemin du fichier CSV à prévisualiser
        """
        try:
            # Détection de l'encodage du fichier
            encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
            delimiter = ','
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        sample = f.read(4096)
                        sniffer = csv.Sniffer()
                        delimiter = sniffer.sniff(sample).delimiter
                        break
                except Exception:
                    continue
            
            # Lecture du CSV avec pandas
            df = pd.read_csv(file_path, delimiter=delimiter, header=None, encoding=encoding)
            
            # Limiter le nombre de lignes
            if len(df) > 200 + self.spinbox_header_start.value() + self.spinbox_header.value():
                df = df.iloc[:(200 + self.spinbox_header_start.value() + self.spinbox_header.value())]
            
            # Récupérer les en-têtes
            headers = []
            header_start = self.spinbox_header_start.value() - 1
            header_rows = self.spinbox_header.value()
            
            for row in range(header_start, header_start + header_rows):
                if row < len(df):
                    headers.append(df.iloc[row].tolist())
            
            # Récupérer les données
            preview_data = []
            for row in range(header_start + header_rows, len(df)):
                preview_data.append(df.iloc[row].tolist())
            
            # Afficher les données
            self.display_preview_tab(headers, preview_data)
            
        except Exception as e:
            raise ValueError(f"Erreur lors de la lecture du fichier CSV: {str(e)}")
    
    def display_preview_tab(self, headers, data):
        """
        Affiche les données dans le tableau de prévisualisation de l'onglet.
        
        Args:
            headers: Données d'en-tête
            data: Données à afficher
        """
        if not headers or not headers[0]:
            return
        
        # Configurer le tableau
        self.preview_table.clear()
        self.preview_table.setRowCount(len(data))
        self.preview_table.setColumnCount(len(headers[-1]))
        
        # Définir les en-têtes des colonnes
        self.preview_table.setHorizontalHeaderLabels([str(h) if h is not None else "" for h in headers[-1]])
        
        # Ajouter les données
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data[:len(headers[-1])]):
                item = QTableWidgetItem(str(value) if value is not None else "")
                self.preview_table.setItem(row_idx, col_idx, item)
        
        # Ajuster la taille des colonnes
        self.preview_table.resizeColumnsToContents()
        
        # Mettre à jour l'étiquette d'information
        if len(data) >= 200:
            self.preview_info_label.setText(self.translate("preview_limited", 200))
        else:
            self.preview_info_label.setText(f"{len(data)} lignes affichées")

    # =====================================================
    # PARTIE 9: MÉTHODES PRINCIPALES
    # =====================================================
    
    def choose_directory(self):
        """Ouvre une boîte de dialogue pour choisir le répertoire de travail."""
        directory = QFileDialog.getExistingDirectory(
            self,
            self.translate("choose_directory"),
            self.directory if self.directory else os.path.expanduser("~")
        )
        
        if directory:
            self.directory = directory
            self.label_directory.setText(directory)
            self.load_files()
            logging.info(f"Répertoire sélectionné: {directory}")
    
    def load_files(self):
        """Charge la liste des fichiers Excel et CSV du répertoire sélectionné."""
        if not self.directory:
            return
        
        self.files = []
        self.list_files.clear()
        self.preview_combo.clear()
        
        # Extensions supportées
        excel_extensions = ['.xlsx', '.xls'] if True else []
        csv_extensions = ['.csv'] if self.checkbox_csv.isChecked() else []
        extensions = excel_extensions + csv_extensions
        
        try:
            for file in os.listdir(self.directory):
                if any(file.lower().endswith(ext) for ext in extensions):
                    self.files.append(file)
                    item = QListWidgetItem(file)
                    self.list_files.addItem(item)
                    self.preview_combo.addItem(file)
            
            logging.info(f"Fichiers chargés: {len(self.files)} fichiers trouvés")
            
            # Mettre à jour le compteur
            self.update_selection_count()
            
            # Précharger la prévisualisation du premier fichier si disponible
            if self.files:
                self.refresh_preview()
                
        except PermissionError:
            QMessageBox.warning(
                self,
                self.translate("error"),
                "Accès au répertoire refusé. Vérifiez vos permissions."
            )
        except Exception as e:
            QMessageBox.warning(
                self,
                self.translate("error"),
                f"Erreur lors du chargement des fichiers: {str(e)}"
            )
    
    def compile_files(self):
        """Lance la compilation des fichiers sélectionnés."""
        selected_files = [item.text() for item in self.list_files.selectedItems()]
        
        if not selected_files:
            QMessageBox.warning(
                self,
                self.translate("warning"),
                self.translate("select_files_message")
            )
            return
        
        # Vérification préliminaire si activée
        if self.verification_enabled:
            compatible_files, incompatible_files = self.verify_files(selected_files)
            
            if incompatible_files:
                # Afficher le rapport de vérification
                dialog = VerificationReportDialog(self, compatible_files, incompatible_files)
                result = dialog.exec()
                
                if result == QDialog.DialogCode.Rejected:
                    return  # L'utilisateur a annulé
                
                if dialog.continue_with_compatible:
                    # Continuer avec seulement les fichiers compatibles
                    selected_files = compatible_files
                    if not selected_files:
                        QMessageBox.information(
                            self,
                            self.translate("info"),
                            "Aucun fichier compatible pour la compilation."
                        )
                        return
        
        # Préparer les paramètres de compilation
        header_start_row = self.spinbox_header_start.value()
        header_rows = self.spinbox_header.value()
        add_filename = self.checkbox_add_filename.isChecked()
        sort_data = self.checkbox_sort_data.isChecked()
        sort_column = self.get_sort_column_index()
        repeat_headers = self.checkbox_repeat_header.isChecked()
        remove_empty_rows = self.checkbox_remove_empty_rows.isChecked()
        remove_duplicates = self.checkbox_remove_duplicates.isChecked()
        
        # Désactiver le bouton et afficher la barre de progression
        self.button_compile.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(len(selected_files))
        self.progress_bar.setValue(0)
        self.status_label.setText(self.translate("compilation_in_progress"))
        
        # Créer et démarrer le worker
        self.compilation_worker = CompilationWorker(
            selected_files,
            self.directory,
            header_start_row,
            header_rows,
            add_filename,
            sort_data,
            sort_column,
            repeat_headers,
            remove_empty_rows,
            remove_duplicates,
            self.date_format
        )
        
        # Connecter les signaux
        self.compilation_worker.progress.connect(self.update_progress)
        self.compilation_worker.error.connect(self.compilation_error)
        self.compilation_worker.finished.connect(self.compilation_finished)
        
        # Démarrer la compilation
        self.compilation_worker.start()
        
        logging.info(f"Début de la compilation de {len(selected_files)} fichiers")
    
    def verify_files(self, file_list):
        """
        Vérifie la compatibilité des fichiers pour la compilation.
        
        Args:
            file_list: Liste des noms de fichiers à vérifier
            
        Returns:
            Tuple[List[str], List[Tuple[str, str]]]: (fichiers_compatibles, fichiers_incompatibles_avec_raisons)
        """
        compatible_files = []
        incompatible_files = []
        
        header_start_row = self.spinbox_header_start.value()
        header_rows = self.spinbox_header.value()
        
        # Créer une barre de progression pour la vérification
        progress_dialog = QProgressDialog(
            "Vérification des fichiers en cours...",
            "Annuler",
            0,
            len(file_list),
            self
        )
        progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
        progress_dialog.setMinimumDuration(0)
        
        for i, file_name in enumerate(file_list):
            if progress_dialog.wasCanceled():
                break
                
            progress_dialog.setValue(i)
            progress_dialog.setLabelText(f"Vérification: {file_name}")
            
            file_path = os.path.join(self.directory, file_name)
            
            try:
                if file_name.lower().endswith(('.xlsx', '.xls')):
                    is_compatible, reason = FileVerification.verify_excel_file(
                        file_path, header_start_row, header_rows
                    )
                elif file_name.lower().endswith('.csv'):
                    is_compatible, reason = FileVerification.verify_csv_file(
                        file_path, header_start_row, header_rows
                    )
                else:
                    is_compatible = False
                    reason = "Format de fichier non supporté"
                
                if is_compatible:
                    compatible_files.append(file_name)
                else:
                    incompatible_files.append((file_name, reason))
                    
            except Exception as e:
                incompatible_files.append((file_name, f"Erreur inattendue: {str(e)}"))
        
        progress_dialog.setValue(len(file_list))
        progress_dialog.close()
        
        return compatible_files, incompatible_files
    
    def show_preview(self):
        """Affiche la boîte de dialogue de prévisualisation des données."""
        selected_files = [item.text() for item in self.list_files.selectedItems()]
        
        if not selected_files:
            QMessageBox.information(
                self,
                self.translate("preview"),
                "Veuillez sélectionner au moins un fichier à prévisualiser."
            )
            return
        
        try:
            dialog = PreviewDialog(
                self,
                self.directory,
                selected_files,
                self.spinbox_header_start.value(),
                self.spinbox_header.value()
            )
            dialog.exec()
        except Exception as e:
            QMessageBox.warning(
                self,
                self.translate("error"),
                f"Erreur lors de la prévisualisation: {str(e)}"
            )
    
    def show_date_format_dialog(self):
        """Affiche la boîte de dialogue de choix du format de date."""
        dialog = DateFormatDialog(self, self.date_format)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.date_format = dialog.get_selected_format()
            
            # Mettre à jour l'onglet format de date
            if self.date_format in self.date_radio_buttons:
                self.date_radio_buttons[self.date_format].setChecked(True)
                if self.date_format == "CUSTOM":
                    self.date_custom_edit.setEnabled(True)
                    self.date_custom_edit.setText(DATE_FORMATS["CUSTOM"]["format"])
                else:
                    self.date_custom_edit.setEnabled(False)
            
            self.update_date_example()
    
    def show_language_dialog(self):
        """Affiche la boîte de dialogue de choix de langue."""
        dialog = LanguageDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_language = dialog.get_selected_language()
            TranslationManager().set_language(new_language)
    
    def show_about_dialog(self):
        """Affiche la boîte de dialogue À propos."""
        QMessageBox.about(
            self,
            self.translate("about"),
            f"""<h2>{self.translate("app_title")}</h2>
            <p>Version 3.0</p>
            <p>Développé par: <b>{self.translate("developer_name")}</b></p>
            <p>{self.translate("developer_title")}</p>
            <p>Email: zimkada@gmail.com</p>
            <p><i>{self.translate("copyright_notice")}</i></p>"""
        )
    
    def show_ip_warning(self):
        """Affiche l'avertissement de propriété intellectuelle au démarrage."""
        dialog = IPWarningDialog(self)
        if dialog.exec() == QDialog.DialogCode.Rejected:
            # L'utilisateur a choisi de quitter
            sys.exit(0)
    
    def save_settings(self):
        """Sauvegarde les paramètres actuels."""
        SettingsManager().save_settings(self)
        QMessageBox.information(
            self,
            self.translate("save_settings"),
            self.translate("settings_saved")
        )
    
    def load_settings(self):
        """Charge les paramètres sauvegardés."""
        if SettingsManager().load_settings(self):
            QMessageBox.information(
                self,
                self.translate("load_settings"),
                self.translate("settings_loaded")
            )
        else:
            QMessageBox.information(
                self,
                self.translate("load_settings"),
                "Aucun paramètre sauvegardé trouvé."
            )

    # =====================================================
    # PARTIE 10: MÉTHODES FINALES
    # =====================================================
    
    def get_sort_column_index(self):
        """
        Obtient l'index de la colonne de tri à partir du texte saisi.
        
        Returns:
            int: Index de la colonne (0-based)
        """
        sort_text = self.lineedit_sort_column.text().strip().upper()
        
        if not sort_text:
            return 0
        
        # Conversion lettre vers index (A=0, B=1, etc.)
        if len(sort_text) == 1 and 'A' <= sort_text <= 'Z':
            return ord(sort_text) - ord('A')
        
        # Conversion numéro vers index (1=0, 2=1, etc.)
        try:
            return max(0, int(sort_text) - 1)
        except ValueError:
            return 0
    
    def update_progress(self, value):
        """
        Met à jour la barre de progression.
        
        Args:
            value: Valeur actuelle de progression
        """
        self.progress_bar.setValue(value)
    
    def compilation_error(self, error_message):
        """
        Gère les erreurs de compilation.
        
        Args:
            error_message: Message d'erreur
        """
        logging.error(f"Erreur de compilation: {error_message}")
        self.status_label.setText(f"Erreur: {error_message}")
    
    def compilation_finished(self, result):
        """
        Méthode appelée à la fin de la compilation.
        
        Args:
            result: Tuple contenant les résultats de la compilation
        """
        preliminary_info, headers, combined_data, merged_cells, successful_files, failed_files = result
        
        # Réactiver l'interface
        self.button_compile.setEnabled(True)
        self.progress_bar.setVisible(False)
        
        if not combined_data or not headers:
            self.status_label.setText(self.translate("no_data"))
            QMessageBox.warning(
                self,
                self.translate("warning"),
                self.translate("no_data_message")
            )
            return
        
        # Générer le nom du fichier de sortie
        output_filename = self.lineedit_output_name.text().strip()
        if not output_filename:
            output_filename = "compilation"
        
        if not output_filename.endswith('.xlsx'):
            output_filename += '.xlsx'
        
        output_path = os.path.join(self.directory, output_filename)
        
        try:
            # Écrire les données dans le fichier Excel
            self.write_excel_file(
                output_path,
                preliminary_info,
                headers,
                combined_data,
                merged_cells
            )
            
            # Afficher le rapport de compilation
            self.show_compilation_report(successful_files, failed_files, output_path)
            
            self.status_label.setText(self.translate("compilation_complete", output_path))
            logging.info(f"Compilation terminée: {output_path}")
            
        except PermissionError:
            self.status_label.setText(self.translate("file_open_error"))
            QMessageBox.warning(
                self,
                self.translate("error"),
                self.translate("file_open_error_message")
            )
        except Exception as e:
            error_msg = str(e)
            self.status_label.setText(self.translate("compilation_failed", error_msg))
            QMessageBox.critical(
                self,
                self.translate("error"),
                self.translate("compilation_failed", error_msg)
            )
            logging.error(f"Erreur lors de l'écriture: {error_msg}")
    
    def write_excel_file(self, output_path, preliminary_info, headers, data, merged_cells):
        """
        Écrit les données compilées dans un fichier Excel.
        
        Args:
            output_path: Chemin du fichier de sortie
            preliminary_info: Informations préliminaires
            headers: En-têtes
            data: Données
            merged_cells: Cellules fusionnées
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Compilation"
        
        current_row = 1
        
        # Écrire les informations préliminaires
        if preliminary_info:
            current_row = ExcelFormatter.write_preliminary_info(ws, preliminary_info)
        
        # Écrire les en-têtes
        if headers:
            header_start_row = current_row
            current_row = ExcelFormatter.write_headers(ws, headers, current_row)
            
            # Appliquer les cellules fusionnées
            if merged_cells:
                ExcelFormatter.apply_merged_cells(
                    ws, merged_cells, 
                    self.spinbox_header_start.value(), 
                    header_start_row
                )
        
        # Écrire les données
        if data:
            ExcelFormatter.write_data(ws, data, current_row, self.date_format)
        
        # Appliquer le formatage
        if self.checkbox_auto_width.isChecked():
            ExcelFormatter.adjust_column_widths(ws)
        
        if self.checkbox_freeze_header.isChecked() and headers:
            freeze_row = header_start_row + len(headers) if preliminary_info else len(headers) + 1
            ExcelFormatter.freeze_header(ws, freeze_row)
        
        # Sauvegarder le fichier
        wb.save(output_path)
    
    def show_compilation_report(self, successful_files, failed_files, output_path):
        """
        Affiche le rapport de compilation.
        
        Args:
            successful_files: Liste des fichiers compilés avec succès
            failed_files: Liste des fichiers échoués
            output_path: Chemin du fichier de sortie
        """
        dialog = CompilationReportDialog(self, successful_files, failed_files, output_path)
        dialog.exec()
    
    def closeEvent(self, event):
        """
        Gère la fermeture de l'application.
        
        Args:
            event: Événement de fermeture
        """
        # Sauvegarder automatiquement les paramètres
        SettingsManager().save_settings(self)
        
        # Arrêter le worker s'il est en cours d'exécution
        if self.compilation_worker and self.compilation_worker.isRunning():
            self.compilation_worker.terminate()
            self.compilation_worker.wait()
        
        # Arrêter le timer de date/heure
        if hasattr(self, 'datetime_timer'):
            self.datetime_timer.stop()
        
        logging.info("Application fermée")
        event.accept()


# =====================================================
# TESTS UNITAIRES
# =====================================================

class TestExcelCompiler(unittest.TestCase):
    """Tests unitaires pour l'application."""
    
    def setUp(self):
        """Configuration des tests."""
        self.app = QApplication([])
        self.compiler = ModernExcelCompilerApp()
    
    def tearDown(self):
        """Nettoyage après les tests."""
        self.compiler.close()
        self.app.quit()
    
    def test_translation_manager(self):
        """Test du gestionnaire de traduction."""
        tm = TranslationManager()
        
        # Test changement de langue
        tm.set_language("en")
        self.assertEqual(tm.current_language, "en")
        self.assertEqual(tm.get_text("app_title"), "Professional Excel Compiler")
        
        tm.set_language("fr")
        self.assertEqual(tm.current_language, "fr")
        self.assertEqual(tm.get_text("app_title"), "Compilateur Excel Professionnel")
    
    def test_file_verification(self):
        """Test de la vérification des fichiers."""
        # Créer un fichier de test temporaire
        import tempfile
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'Test Header'
            ws['A2'] = 'Test Data'
            wb.save(temp_file.name)
            
            # Test de vérification
            is_compatible, reason = FileVerification.verify_excel_file(
                temp_file.name, 1, 1
            )
            self.assertTrue(is_compatible)
            
        # Nettoyer
        os.unlink(temp_file.name)
    
    def test_date_formats(self):
        """Test des formats de date."""
        self.assertIn("STANDARD", DATE_FORMATS)
        self.assertIn("FRENCH", DATE_FORMATS)
        self.assertIn("format", DATE_FORMATS["FRENCH"])  
        self.assertIn("excel_format", DATE_FORMATS["FRENCH"])
    
    def test_sort_column_conversion(self):
        """Test de la conversion des colonnes de tri."""
        self.compiler.lineedit_sort_column.setText("A")
        self.assertEqual(self.compiler.get_sort_column_index(), 0)
        
        self.compiler.lineedit_sort_column.setText("B")
        self.assertEqual(self.compiler.get_sort_column_index(), 1)
        
        self.compiler.lineedit_sort_column.setText("1")
        self.assertEqual(self.compiler.get_sort_column_index(), 0)


# =====================================================
# FONCTION PRINCIPALE
# =====================================================

def main():
    """Fonction principal pour lancer l'application."""
    try:
        # Configuration de l'application
        app = QApplication(sys.argv)
        app.setApplicationName("Excel Compiler")
        app.setApplicationVersion("3.0")
        app.setOrganizationName("GOUNOU N'GOBI")
        app.setOrganizationDomain("zimkada@gmail.com")
        
        # Définir l'icône de l'application si disponible
        if os.path.exists("icon.jpg"):
            app.setWindowIcon(QIcon("icon.jpg"))
        
        # Style de l'application
        app.setStyle('Fusion')
        
        # Palette de couleurs personnalisée
        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window, QColor(240, 240, 240))
        palette.setColor(QPalette.ColorRole.WindowText, QColor(0, 0, 0))
        palette.setColor(QPalette.ColorRole.Base, QColor(255, 255, 255))
        palette.setColor(QPalette.ColorRole.AlternateBase, QColor(245, 245, 245))
        palette.setColor(QPalette.ColorRole.ToolTipBase, QColor(255, 255, 220))
        palette.setColor(QPalette.ColorRole.ToolTipText, QColor(0, 0, 0))
        palette.setColor(QPalette.ColorRole.Text, QColor(0, 0, 0))
        palette.setColor(QPalette.ColorRole.Button, QColor(240, 240, 240))
        palette.setColor(QPalette.ColorRole.ButtonText, QColor(0, 0, 0))
        palette.setColor(QPalette.ColorRole.BrightText, QColor(255, 0, 0))
        palette.setColor(QPalette.ColorRole.Link, QColor(42, 130, 218))
        palette.setColor(QPalette.ColorRole.Highlight, QColor(42, 130, 218))
        palette.setColor(QPalette.ColorRole.HighlightedText, QColor(0, 0, 0))
        app.setPalette(palette)
        
        # Créer et afficher la fenêtre principale
        compiler = ModernExcelCompilerApp()
        compiler.show()
        
        # Configuration du logging pour capturer les erreurs non gérées
        def handle_exception(exc_type, exc_value, exc_traceback):
            if issubclass(exc_type, KeyboardInterrupt):
                sys.__excepthook__(exc_type, exc_value, exc_traceback)
                return
            
            logging.critical("Exception non gérée", exc_info=(exc_type, exc_value, exc_traceback))
            QMessageBox.critical(
                None,
                "Erreur Critique",
                f"Une erreur inattendue s'est produite:\n{exc_type.__name__}: {exc_value}"
            )
        
        sys.excepthook = handle_exception
        
        # Démarrer la boucle d'événements
        sys.exit(app.exec())
        
    except Exception as e:
        logging.critical(f"Erreur critique au démarrage: {str(e)}")
        logging.critical(traceback.format_exc())
        
        # Créer une application minimale pour afficher l'erreur
        if 'app' not in locals():
            app = QApplication(sys.argv)
        
        QMessageBox.critical(
            None,
            "Erreur de Démarrage",
            f"Impossible de démarrer l'application:\n{str(e)}\n\nConsultez le fichier de log pour plus de détails."
        )
        sys.exit(1)


if __name__ == "__main__":
    main()

    


